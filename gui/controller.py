from __future__ import annotations
import datetime, random, urllib.parse
from pathlib import Path
from typing  import Dict, List, Tuple

import openpyxl
from PySide6.QtGui       import QDesktopServices
from PySide6.QtCore      import QUrl

from ..settings          import GHIBLI_SHEET_PATH, SPREADSHEET_ID
from ..utils             import normalize, fuzzy_match, log_debug, open_url_host_browser
from ..metadata.movie_night_db  import DB, Movie
from ..metadata.api_clients.tmdb_client import tmdb_client
from ..metadata.api_clients.omdb_client import omdb_client     
from ..metadata import locate_trailer        # you already have this helper
from ..movie_api import sheets_xlsx


# type alias for what we return to the GUI
GenerateResult = Tuple[List[str], Dict[str, str | None]]

#This should only pull from spreadsheet_themes in the DB
def generate_movies(sheet_name_input: str, attendee_count: int) -> GenerateResult:
    """
    Core “pick movies” routine, decoupled from Qt slots.

    Returns:
        (chosen_titles, {title: trailer_url_or_None})

    Raises:
        ValueError with user-friendly .args[0] for any validation issue.
    """
    # ---- validation -------------------------------------------------
    if attendee_count <= 0:
        raise ValueError("Enter a positive attendee count.")

    if not sheet_name_input.strip():
        raise ValueError("Sheet name?")

    if not Path(GHIBLI_SHEET_PATH).exists():
        raise ValueError("Run ‘Update data’ first.")

    # ---- open workbook & resolve sheet ------------------------------
    wb   = openpyxl.load_workbook(GHIBLI_SHEET_PATH, read_only=True)
    norm = {normalize(n): n for n in wb.sheetnames}
    chosen_sheet = (
        norm.get(normalize(sheet_name_input))
        or norm.get(fuzzy_match(normalize(sheet_name_input), list(norm)))
    )
    if not chosen_sheet:
        raise ValueError("Sheet not found.")

    titles = [
        r[0] for r in wb[chosen_sheet].iter_rows(min_row=1, max_col=1, values_only=True)
        if r[0]
    ]
    if attendee_count + 1 > len(titles):
        raise ValueError("Not enough movies on that sheet.")

    # ---- pick and locate trailers -----------------------------------
    picks = random.sample(titles, attendee_count + 1)
    trailer_map: Dict[str, str | None] = {}

    for title in picks:
        # locate_trailer uses DB + TMDb + OMDb under the hood
        url, _src, _ = locate_trailer(chosen_sheet, title)
        trailer_map[title] = url

    # ---- build & auto-open YT playlist ------------------------------
    ids = [
        url.split("v=")[-1].split("&")[0]
        for url in trailer_map.values()
        if url and "watch?v=" in url
    ]
    if ids:
        playlist_link = (
            "https://www.youtube.com/watch_videos"
            f"?video_ids={','.join(ids)}"
            f"&title={urllib.parse.quote_plus(f'Movie Night {datetime.date.today()}')}"
            "&feature=share"
        )
        open_url_host_browser(playlist_link)

    return picks, trailer_map
# ------------------------------------------------------------------
def add_remove_movie(current: list[str], pool: list[str], delta: int) -> list[str]:
    """
    +1 ⇒ add a random title from *pool* not already in list
    –1 ⇒ remove one random title (but keep ≥1)
    Returns new list.
    """
    lst = current.copy()
    if delta > 0:
        choice = random.choice([t for t in pool if t not in lst])
        lst.append(choice)
    elif len(lst) > 1:
        lst.pop(random.randrange(len(lst)))
    return lst

def update_data() -> None:
    """
    • Sync spreadsheet rows with DB
    • Fill trailer + full metadata:
        TMDb first → OMDb fills remaining blanks
    • Re-compute ratings, trends, combined_score for touched movies
    """
    # 1) ensure local XLSX
    sheets_xlsx.download_spreadsheet_as_xlsx(SPREADSHEET_ID, GHIBLI_SHEET_PATH)

    touched_ids: set[int] = set()      # movie-ids we’ll recalc at the end

    for sheet in sheets_xlsx.get_non_green_tabs(SPREADSHEET_ID):

        theme_id = DB._ensure_spreadsheet_theme(sheet)

        for title in sheets_xlsx.get_movie_titles_from_sheet(GHIBLI_SHEET_PATH, sheet):

            mid  = DB.get_movie_id_by_title(title) or DB.add_movie({"title": title})
            DB.link_movie_to_sheet_theme(mid, theme_id)
            DB.upsert_trailer_url(sheet, title, None)

            # ---------------- current state ------------------------
            mrow = DB.get_movie(mid)
            trow = DB.get_trailer_url(sheet, title)

            # ---------------- TMDb primary call --------------------
            tm_meta = tmdb.fetch_metadata(title=title) or {}
            mf      = tm_meta.get("movie_fields", {})

            # trailer
            if not trow and mf.get("youtube_link"):
                DB.upsert_trailer_url(sheet, title, mf["youtube_link"])

            # numeric & string fields
            for fld in ("year", "release_window", "rating_cert", "origin_country",
                        "duration_seconds", "box_office_actual", "franchise"):
                if DB.is_movie_field_missing(mid, fld) and mf.get(fld):
                    DB.update_movie_field(mid, fld, mf[fld])

            # genres
            for g in tm_meta.get("genres", []):
                DB.link_movie_genre(mid, g)

            # ---------------- OMDb fallback (only if still missing) ----
            # runtime
            if DB.is_movie_field_missing(mid, "duration_seconds"):
                rt = omdb_client.get_runtime_seconds(title=title)
                if rt:
                     DB.update_movie_field(mid, "duration_seconds", rt)

            # box-office actual
            if DB.is_movie_field_missing(mid, "box_office_actual"):
                bo = omdb_client.get_boxoffice(title=title)
                if bo:
                    DB.update_movie_field(mid, "box_office_actual", bo)

            touched_ids.add(mid)

    DB.conn.commit()

    # 2) compute ratings + trends for every touched movie
    for mid in touched_ids:
        Movie.from_id(mid).update_scores_and_trends()

    log_debug(f"Update data complete ({len(touched_ids)} movies refreshed).")