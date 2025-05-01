import random
import datetime
import urllib.parse
import openpyxl
from pathlib import Path

from PySide6.QtCore    import QUrl
from PySide6.QtGui     import QDesktopServices
from PySide6.QtWidgets import QMessageBox

from ..settings        import GHIBLI_SHEET_PATH, SPREADSHEET_ID
from ..movie_api.sheets_xlsx   import download_spreadsheet_as_xlsx, get_non_green_tabs
from ..temp_modules.json_functions   import (
    build_master_cache_from_all_json,
    ensure_url_json_exists)
from ..temp_modules.json_cache import fill_missing_urls_in_json_with_cache
from ..movie_api.sheets_xlsx import get_movie_titles_from_sheet
from ..movie_api.youtube     import locate_trailer
from ..utils           import normalize, fuzzy_match, log_debug

def generate_movies(
    attendee_text: str,
    sheet_name_text: str,
    parent_widget
) -> tuple[list[str], dict[str, str]]:
    """
    1) Validate inputs,
    2) Randomly pick movies (count+1),
    3) Lookup trailers,
    4) Fire off a YouTube watch_videos link,
    5) Return (selected_titles, {title: trailer_url}).
    Raises ValueError on any validation failure.
    """
    # -- Parse & validate attendee count --
    try:
        attendee_count = int(attendee_text.strip())
        if attendee_count <= 0:
            raise ValueError("Enter a positive attendee count.")
    except Exception:
        raise ValueError("Enter a positive attendee count.")
    # -- Validate sheet name --
    sheet_raw = sheet_name_text.strip()
    if not sheet_raw:
        raise ValueError("Please enter a sheet name.")
    if not GHIBLI_SHEET_PATH.exists():
        raise ValueError("Spreadsheet not found. Run Update first.")
    # -- Load and fuzzy‐find the worksheet --
    wb = openpyxl.load_workbook(GHIBLI_SHEET_PATH, read_only=True)
    normalized_map = { normalize(name): name for name in wb.sheetnames }
    lookup_key = normalize(sheet_raw)
    real_sheet = normalized_map.get(lookup_key) or normalized_map.get(
        fuzzy_match(lookup_key, list(normalized_map))
    )
    if not real_sheet:
        raise ValueError(f"Worksheet “{sheet_raw}” not found.")
    # -- Extract titles --
    titles = [
        row[0]
        for row in wb[real_sheet].iter_rows(min_row=1, max_col=1, values_only=True)
        if row[0]
    ]
    if attendee_count + 1 > len(titles):
        raise ValueError("Not enough movies in that sheet.")
    # -- Sample & lookup trailers --
    selected = random.sample(titles, attendee_count + 1)
    trailer_map = { title: locate_trailer(real_sheet, title)[0] for title in selected }
    # -- Build & open watch_videos link if any IDs found --
    video_ids = [
        url.split("v=")[-1].split("&")[0]
        for url in trailer_map.values() if url and "watch?v=" in url
    ]
    if video_ids:
        csv_ids = ",".join(video_ids)
        playlist_name = urllib.parse.quote_plus(f"Movie Night {datetime.date.today()}")
        watch_url = (
            f"https://www.youtube.com/watch_videos"
            f"?video_ids={csv_ids}&title={playlist_name}&feature=share"
        )
        QDesktopServices.openUrl(QUrl(watch_url))
    return selected, trailer_map

def update_trailer_urls() -> None:
    """
    1) Download latest spreadsheet,
    2) Find all non-green tabs,
    3) Build a master cache from existing JSON files,
    4) For each tab, ensure JSON exists, fill missing URLs.
    """
    # 1) Download fresh XLSX
    download_spreadsheet_as_xlsx(SPREADSHEET_ID, GHIBLI_SHEET_PATH)
    # 2) Get non-green tabs
    tab_list = get_non_green_tabs(SPREADSHEET_ID)
    if not tab_list:
        log_debug("No non-green sheets to process.")
        return
    # 3) Build master cache of all existing JSON files
    master_cache = build_master_cache_from_all_json()
    # 4) Process each sheet
    for sheet_title in tab_list:
        try:
            movie_list = get_movie_titles_from_sheet(GHIBLI_SHEET_PATH, sheet_title)
        except Exception as exc:
            log_debug(f"Error reading sheet {sheet_title}: {exc}")
            continue
        if not movie_list:
            log_debug(f"No movies found in sheet '{sheet_title}'.")
            continue
        json_file = ensure_url_json_exists(sheet_title)
        fill_missing_urls_in_json_with_cache(json_file, movie_list, master_cache)
    log_debug("Completed update of trailer URLs.")
