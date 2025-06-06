import io
from pathlib import Path
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from movieNight.settings import GOOGLE_SERVICE_ACCOUNT_FILE, DRIVE_SCOPES, SHEETS_SCOPES
from movieNight.utils import log_debug
load_dotenv(GOOGLE_SERVICE_ACCOUNT_FILE.parent / "secret.env")
import openpyxl

def get_drive_service():
    """Authenticate & return a Google Drive service client."""
    creds = Credentials.from_service_account_file(
        GOOGLE_SERVICE_ACCOUNT_FILE, scopes=DRIVE_SCOPES
    )
    return build("drive", "v3", credentials=creds)

def get_sheets_service():
    """Authenticate & return a Google Sheets service client."""
    creds = Credentials.from_service_account_file(
        GOOGLE_SERVICE_ACCOUNT_FILE, scopes=SHEETS_SCOPES
    )
    return build("sheets", "v4", credentials=creds)

def download_spreadsheet_as_xlsx(spreadsheet_id: str, output_path: Path) -> None:
    """
    Export the given Google Sheet to XLSX and save it at output_path.
    Overwrites any existing file.
    """
    drive_svc = get_drive_service()
    export_req = drive_svc.files().export_media(
        fileId=spreadsheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, export_req)
    done = False
    while not done:
        status, done = downloader.next_chunk()
        if status:
            print(f"Download progress: {int(status.progress() * 100)}%")
    output_path.write_bytes(fh.getvalue())

def get_non_green_tabs(spreadsheet_id: str) -> list[str]:
    """
    Return the titles of all sheets whose tab color is NOT 'green'
    (defined as green≥0.8, red<0.2, blue<0.2).
    """
    sheets_svc = get_sheets_service()
    metadata = sheets_svc.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    result: list[str] = []
    for sheet in metadata.get("sheets", []):
        props = sheet.get("properties", {})
        title = props.get("title", "")
        color = props.get("tabColor", {})
        if not color or not (color.get("green", 0) >= 0.8
                             and color.get("red", 0) < 0.2
                             and color.get("blue", 0) < 0.2):
            result.append(title)
    return result
def get_movie_titles_from_sheet(
    excel_path: Path,
    sheet_name: str
) -> list[str]:
    """
    Open the .xlsx at `excel_path`, look for `sheet_name`, and
    return all non-empty values from column A as a list of strings.
    Logs an error and returns [] if the sheet is missing.
    """
    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
    except Exception as e:
        log_debug(f"[ERROR] Failed to load workbook {excel_path}: {e}")
        return []

    if sheet_name not in workbook.sheetnames:
        log_debug(f"[ERROR] Sheet '{sheet_name}' not found in {excel_path.name}.")
        return []

    sheet = workbook[sheet_name]
    titles: list[str] = []
    for (cell_value,) in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        if cell_value and str(cell_value).strip():
            titles.append(str(cell_value).strip())

    return titles