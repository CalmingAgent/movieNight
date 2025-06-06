import os
import sys
import json
import re
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv
import datetime
from yt_dlp import YoutubeDL

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google.oauth2.service_account import Credentials

# ----------------- CONFIGURATION ----------------- #
BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / "secret.env"

load_dotenv(ENV_PATH)

TMDB_API_KEY = os.getenv("TMDB_API_KEY")
YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY")
SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")  # to read tab colors

if not TMDB_API_KEY:
    raise EnvironmentError("Missing 'TMDB_API_KEY' in secret.env or environment variables.")
if not YOUTUBE_API_KEY:
    raise EnvironmentError("Missing 'YOUTUBE_API_KEY' in secret.env or environment variables.")
if not SPREADSHEET_ID:
    raise EnvironmentError("Missing 'SPREADSHEET_ID' in secret.env or environment variables.")

GHIB_FILE = BASE_DIR / "ghib.xlsx"
TRAILERS_DIR = BASE_DIR / "Video_Trailers"
TRAILERS_DIR.mkdir(parents=True, exist_ok=True)

LOG_FILE = BASE_DIR / "trailer_debug.log"

GOOGLE_SERVICE_ACCOUNT_FILE = BASE_DIR / "service_secret.json"
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
SHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]

YOUTUBE_MAXED_OUT = False  # if True, skip further YT lookups

# --------------- LOGGING & UTILS --------------- #
def log_debug(message: str) -> None:
    timestamp = datetime.datetime.now().isoformat(timespec='seconds')
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {message}\n")

def sanitize_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', '', name.strip())

def normalize_title(title: str) -> str:
    return re.sub(r'[^a-z0-9]', '', title.lower().strip())

def print_progress_bar(iteration: int, total: int, prefix: str="", suffix: str="", length: int=40):
    if total <= 0:
        return
    fraction = iteration / float(total)
    filled_length = int(length * fraction)
    bar = "█" * filled_length + "-" * (length - filled_length)
    percent = round(100 * fraction, 1)
    print(f"\r{prefix} |{bar}| {percent}% {suffix}", end="\r")
    if iteration >= total:
        print()

# ---------------- DRIVE + SHEETS API ---------------- #
def get_drive_service():
    creds = Credentials.from_service_account_file(GOOGLE_SERVICE_ACCOUNT_FILE, scopes=DRIVE_SCOPES)
    return build("drive", "v3", credentials=creds)

def get_sheets_service():
    creds = Credentials.from_service_account_file(GOOGLE_SERVICE_ACCOUNT_FILE, scopes=SHEETS_SCOPES)
    return build("sheets", "v4", credentials=creds)

def download_spreadsheet_as_xlsx(spreadsheet_id: str, out_file: Path):
    """
    Export Google Sheet as .xlsx, overwriting out_file.
    """
    drive_service = get_drive_service()
    request = drive_service.files().export_media(
        fileId=spreadsheet_id,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    from googleapiclient.http import MediaIoBaseDownload
    import io

    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
        if status:
            log_debug(f"Download {int(status.progress() * 100)}%.")
    out_file.write_bytes(fh.getvalue())
    log_debug(f"Downloaded spreadsheet to {out_file}")

def get_non_green_tabs(spreadsheet_id: str):
    """
    Use Sheets API to read tabColor. Skip if 'green'.
    We'll define 'green' as (green >=0.8) and (red<0.2) and (blue<0.2).
    """
    sheets_service = get_sheets_service()
    spreadsheet = sheets_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheets_data = spreadsheet.get("sheets", [])
    result = []
    for sheet in sheets_data:
        props = sheet.get("properties", {})
        title = props.get("title", "")
        tab_color = props.get("tabColor")
        if not tab_color:
            # no color => treat it as 'non-green'
            result.append(title)
            continue
        red   = tab_color.get("red", 0)
        green = tab_color.get("green", 0)
        blue  = tab_color.get("blue", 0)

        if (green >= 0.8) and (red < 0.2) and (blue < 0.2):
            log_debug(f"Skipping sheet '{title}' because tab is green.")
        else:
            result.append(title)

    return result

# ---------------- XLSX READING ---------------- #
def get_all_sheet_names_local(xlsx_file: Path):
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    return wb.sheetnames

def fetch_movie_list_local(xlsx_file: Path, sheet_name: str):
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    if sheet_name not in wb.sheetnames:
        log_debug(f"[ERROR] Sheet '{sheet_name}' not found in {xlsx_file.name}.")
        return []
    sheet = wb[sheet_name]
    movies = []
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).strip():
            movies.append(str(val).strip())
    return movies

# ---------------- TMDB + YOUTUBE ---------------- #
def tmdb_find_trailer(movie_title: str):
    global YOUTUBE_MAXED_OUT
    try:
        norm_query = normalize_title(movie_title)
        all_results = []
        for page in [1, 2]:
            search_url = "https://api.themoviedb.org/3/search/movie"
            params = {
                "api_key": TMDB_API_KEY,
                "query": movie_title,
                "page": page
            }
            log_debug(f"[TMDB] Searching page={page} for '{movie_title}'")
            resp = requests.get(search_url, params=params, timeout=10)
            if resp.status_code == 429:
                log_debug("[TMDB] 429 => daily limit. Exiting.")
                print("TMDB daily quota limit reached. Exiting.")
                sys.exit(1)

            data = resp.json()
            page_results = data.get("results", [])
            if not page_results:
                break
            all_results.extend(page_results)

            total_pages = data.get("total_pages", 1)
            if page >= total_pages:
                break

        log_debug(f"[TMDB] Found {len(all_results)} total results for '{movie_title}' (p1-2).")
        if not all_results:
            log_debug(f"[TMDB] No results => fallback to YT for '{movie_title}'")
            return None

        matched = []
        for r in all_results:
            rtitle = r.get("title", "")
            if normalize_title(rtitle) == norm_query:
                matched.append(r)

        log_debug(f"[TMDB] {len(matched)} EXACT matches for '{movie_title}'")
        if len(matched) == 0 or len(matched) >= 3:
            log_debug(f"[TMDB] Fallback to YT => 0 or >=3 exact matches.")
            return None
        
       
        
            
        chosen = matched[0]
        movie_id = chosen["id"]
        log_debug(f"[TMDB] Using ID={movie_id} for '{chosen.get('title')}'")

        # fetch /videos
        videos_url = f"https://api.themoviedb.org/3/movie/{movie_id}/videos"
        vid_params = {"api_key": TMDB_API_KEY}
        vids_resp = requests.get(videos_url, params=vid_params, timeout=10)
        if vids_resp.status_code == 429:
            log_debug("[TMDB] 429 => daily limit while fetching /videos. Exiting.")
            print("TMDB daily quota limit reached. Exiting.")
            sys.exit(1)

        max_video_duraction = 50
        youtubeLink = "https://www.youtube.com/watch?v="
        vids_data = vids_resp.json().get("results", [])
        for vid in vids_data:
            if vid.get("site") == "YouTube" and vid.get("type") in ("Trailer"):
                key = vid.get("key")
                if key:
                    YTFullLink = youtubeLink + key
                    if get_video_duration_sec(YTFullLink) <= max_video_duraction:
                        log_debug(f"[TMDB] Fallback to YT, {movie_title} video too long, <=50 secs {get_video_duration_sec(matched)}")
                        return None
                    else:
                        log_debug(f"[TMDB] Found YT key={key} for '{movie_title}'")
                        return f"https://www.youtube.com/watch?v={key}"

        log_debug(f"[TMDB] No YT trailer in /videos => fallback to YT for '{movie_title}'")
        return None
    except Exception as e:
        log_debug(f"[ERROR] TMDB search error for '{movie_title}': {e}")
        return None

def youtube_api_search(query: str):
    global YOUTUBE_MAXED_OUT
    if YOUTUBE_MAXED_OUT:
        log_debug(f"[YouTube] Skipping '{query}' => YOUTUBE_MAXED_OUT=True")
        return None

    try:
        log_debug(f"[YouTube] Searching for '{query}'")
        search_url = "https://www.googleapis.com/youtube/v3/search"
        params = {
            "part": "snippet",
            "q": query,
            "key": YOUTUBE_API_KEY,
            "videoDuration": "short",
            "maxResults": 1,
            "type": "video"
        }
        resp = requests.get(search_url, params=params, timeout=10)
        if resp.status_code == 403:
            error_data = resp.json().get("error", {})
            reason_list = error_data.get("errors", [])
            for er in reason_list:
                reason = er.get("reason", "")
                if reason in ("quotaExceeded", "dailyLimitExceeded"):
                    log_debug("[YouTube] Quota exceeded => YOUTUBE_MAXED_OUT=True.")
                    print("YouTube daily quota limit reached. Skipping further YT lookups.")
                    YOUTUBE_MAXED_OUT = True
                    return None
            log_debug(f"[YouTube] 403 error but not recognized reason => none.")
            return None

        data = resp.json()
        items = data.get("items", [])
        if items:
            video_id = items[0]["id"]["videoId"]
            video_title = items[0]["snippet"]["title"]
            log_debug(f"[YouTube] Found video_id={video_id} => '{video_title}' for '{query}'")
            return (f"https://www.youtube.com/watch?v={video_id}", video_title)
        else:
            log_debug(f"[YouTube] No items found for '{query}'")
            return None
    except Exception as e:
        log_debug(f"[ERROR] YouTube API error '{query}': {e}")
        return None
    
def get_video_duration_sec(video_url: str) -> int | None:
    """Return length in seconds, or None on failure (no Google API key needed)."""
    ydl_opts = {"quiet": True, "skip_download": True}
    with YoutubeDL(ydl_opts) as ydl:
        try:
            info = ydl.extract_info(video_url, download=False)
            return info.get("duration")
        except Exception as exc:
            print("yt-dlp error:", exc)
            return None
        
def find_trailer_fallback_cache(movie_title: str, master_cache: dict):
    norm = normalize_title(movie_title)
    if norm in master_cache and master_cache[norm]:
        log_debug(f"[CACHE] Using cached URL for '{movie_title}' => {master_cache[norm]}")
        return master_cache[norm]

    # TMDB
    tmdb_url = tmdb_find_trailer(movie_title)
    if tmdb_url:
        master_cache[norm] = tmdb_url
        log_debug(f"[CACHE] Storing TMDB result => {tmdb_url}")
        return tmdb_url

    # fallback => YouTube
    yt_query = movie_title + " official hd trailer"
    yt_result = youtube_api_search(yt_query)
    if yt_result:
        video_url, _ = yt_result
        master_cache[norm] = video_url
        log_debug(f"[CACHE] Storing YT result => {video_url}")
        return video_url

    log_debug(f"[FALLBACK] No trailer for '{movie_title}' from TMDB or YouTube.")
    return None

# -------------- MASTER CACHE -------------- #
def build_master_cache_from_all_json() -> dict:
    master_cache = {}
    for json_file in TRAILERS_DIR.glob("*.json"):
        try:
            text = json_file.read_text(encoding="utf-8")
            data = json.loads(text)
        except json.JSONDecodeError:
            data = {}
        for title, url in data.items():
            if url:
                norm = normalize_title(title)
                if norm not in master_cache:
                    master_cache[norm] = url
    return master_cache

# -------------- JSON HELPERS -------------- #
def load_json_dict(json_path: Path):
    try:
        return json.loads(json_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}

def write_json_dict(json_path: Path, data: dict):
    lines = ["{"]
    keys = list(data.keys())
    for i, key in enumerate(keys):
        comma = "," if i < len(keys) - 1 else ""
        val = data[key] or ""
        val_escaped = val.replace('"', '\\"')
        lines.append(f'  "{key}": "{val_escaped}"{comma}')
    lines.append("}")
    json_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

def ensure_url_json_exists(sheet_name: str) -> Path:
    safe_sheet_name = sanitize_filename(sheet_name).replace(" ", "")
    json_file = TRAILERS_DIR / f"{safe_sheet_name}Urls.json"
    if not json_file.exists():
        json_file.write_text("{}", encoding="utf-8")
        log_debug(f"[INFO] Created file: {json_file}")
    return json_file

# -------------- FILL MISSING URLS -------------- #
def fill_missing_urls_in_json_with_cache(json_file: Path, movies: list, master_cache: dict):
    data = load_json_dict(json_file)
    for m in movies:
        if m not in data:
            data[m] = ""

    missing_entries = [m for m in movies if not data[m]]
    total_missing = len(missing_entries)
    if total_missing == 0:
        log_debug(f"[INFO] No missing URLs for {json_file.name}.")
        print(f"No missing URLs in {json_file.name} — skipping searches.")
        return

    print(f"\nFilling missing URLs in {json_file.name}: {total_missing} to search.")
    updated = False

    for i, movie_title in enumerate(missing_entries, start=1):
        print_progress_bar(i - 1, total_missing, prefix="Progress", suffix="Complete")
        trailer_url = find_trailer_fallback_cache(movie_title, master_cache)
        if trailer_url:
            data[movie_title] = trailer_url
            updated = True
            log_debug(f"[INFO] Filled '{movie_title}' => {trailer_url}")
        else:
            log_debug(f"[INFO] No trailer found for '{movie_title}'")

    print_progress_bar(total_missing, total_missing, prefix="Progress", suffix="Complete")

    if updated:
        write_json_dict(json_file, data)
        log_debug(f"[INFO] Updated JSON: {json_file}")
        print(f"Completed filling missing trailers in {json_file.name}")
    else:
        print(f"No valid trailer links found for any missing titles in {json_file.name}")


def fill_missing_urls_for_non_green_sheets():
    """
    1) Download .xlsx
    2) Read remote sheet metadata, skip green tabs
    3) For each non-green tab => fill JSON
    """
    # 1) Download
    download_spreadsheet_as_xlsx(SPREADSHEET_ID, GHIB_FILE)

    # 2) Which tabs are non-green
    non_green = get_non_green_tabs(SPREADSHEET_ID)
    if not non_green:
        print("No non-green sheets to process!")
        return

    # 3) Master cache
    master_cache = build_master_cache_from_all_json()

    # 4) For each non-green sheet, fill JSON
    wb = openpyxl.load_workbook(GHIB_FILE, read_only=True)
    for sheet_name in non_green:
        if sheet_name not in wb.sheetnames:
            log_debug(f"[INFO] Non-green sheet '{sheet_name}' not found in local xlsx (maybe renamed?).")
            continue

        movies = []
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
            val = row[0]
            if val and str(val).strip():
                movies.append(str(val).strip())

        if not movies:
            log_debug(f"[INFO] Sheet '{sheet_name}' has no movies in col A.")
            continue

        json_file = ensure_url_json_exists(sheet_name)
        fill_missing_urls_in_json_with_cache(json_file, movies, master_cache)

    print("\nAll non-green sheets processed! Check 'trailer_debug.log' for logs.")

# -------------- MAIN -------------- #
if __name__ == "__main__":
    fill_missing_urls_for_non_green_sheets()
    print("Done!")