import os
import re
import io
import json
import random
import subprocess
import sys
import tkinter as tk
from tkinter import messagebox
from pathlib import Path
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pickle
from PIL import Image, ImageTk
import datetime
from typing import Optional, List
import requests
import openpyxl

# Additional import to download file from Drive
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from difflib import get_close_matches

# --------------------- CONFIGURATION --------------------- #
BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / "secret.env"
LOG_FILE = BASE_DIR / "trailer_debug.log"

# JSON and local XLSX storage
TRAILERS_DIR = BASE_DIR / "Video_Trailers"
NUMBERS_DIR = BASE_DIR / "Numbers"
GHIB_FILE = BASE_DIR / "ghib.xlsx"  # We'll download the spreadsheet to this file each time

# Service account + client secrets
GOOGLE_SERVICE_ACCOUNT_FILE = BASE_DIR / "service_secret.json"
CLIENT_SECRET_FILE = BASE_DIR / "client_secret.json"
YOUTUBE_TOKEN_FILE = BASE_DIR / "youtube_token.pickle"

# Quota-limited Google Drive scope (read spreadsheet as XLSX)
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

# YouTube scope for user-level actions
YOUTUBE_SCOPES = ["https://www.googleapis.com/auth/youtube"]

# For searching YouTube
YOUTUBE_SEARCH_URL = "https://www.googleapis.com/youtube/v3/search"

# -------------DARK MODE THEME COLORS --------------------- #
BACKGROUND_COLOR = "#2e2e2e"
FOREGROUND_COLOR = "#e0e0e0"
BUTTON_COLOR = "#444444"
HIGHLIGHT_COLOR = "#5c5c5c"
ERROR_COLOR = "#ff4c4c"
FALLBACK_COLOR = "#ffa500"

# ----------------- ENVIRONMENT LOADING ------------------- #
load_dotenv(ENV_PATH)
SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")   # The ID of your Google Sheet file
YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY")

if not SPREADSHEET_ID:
    raise EnvironmentError("Missing Spreadsheet ID in secret.env")

if not YOUTUBE_API_KEY:
    raise EnvironmentError("Missing YOUTUBE_API_KEY in secret.env")

# ---------------- UTILS ---------------- #
def log_debug(message: str) -> None:
    """Log debug messages to a file for troubleshooting (cross-platform)."""
    with open(LOG_FILE, "a", encoding="utf-8") as log_file:
        log_file.write(f"{message}\n")


def sanitize_filename(name: str) -> str:
    """Remove characters invalid for filenames; strip leading/trailing spaces."""
    return re.sub(r'[<>:"/\\|?*-]', '', name.strip())


def normalize(text: str) -> str:
    """Lowercase the string and remove all non-alphanumeric characters."""
    return re.sub(r'[^a-z0-9]', '', text.strip().lower())


def fuzzy_search(target: str, candidates: List[str], cutoff=0.8) -> Optional[str]:
    """
    Return the best fuzzy match for 'target' within 'candidates'.
    By default, cutoff=0.8 for an 80% match requirement.
    """
    matches = get_close_matches(target, candidates, n=1, cutoff=cutoff)
    return matches[0] if matches else None

def open_in_windows_default(url: str):
    subprocess.run(["wslview", url])


# ---------------- YOUTUBE TRAILER SEARCH ---------------- #
def youtube_api_search(query: str) -> Optional[tuple]:
    """
    Use the public YouTube Search API to find a single short video by 'query'.
    Return (video_url, video_title) if successful, else None.
    """
    params = {
        "part": "snippet",
        "q": query,
        "key": YOUTUBE_API_KEY,
        "videoDuration": "short",
        "maxResults": 1,
        "type": "video"
    }
    try:
        response = requests.get(YOUTUBE_SEARCH_URL, params=params)
        data = response.json()
        if "items" in data and data["items"]:
            video_id = data["items"][0]["id"]["videoId"]
            video_title = data["items"][0]["snippet"]["title"]
            return f"https://www.youtube.com/watch?v={video_id}", video_title
    except Exception as e:
        log_debug(f"[ERROR] YouTube API search failed: {e}")
    return None


# ---------------- DRIVE (DOWNLOAD SHEET AS XLSX) ---------------- #
def get_drive_service():
    """
    Return an authorized Drive API client using a service account
    with 'drive.readonly' scope. The service account must have read access to the file.
    """
    creds = Credentials.from_service_account_file(
        GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=DRIVE_SCOPES
    )
    return build("drive", "v3", credentials=creds)


def download_spreadsheet_as_xlsx(spreadsheet_id: str, out_file: Path):
    """
    Export the Google Sheet as an .xlsx (Excel) file and save it to 'out_file'.
    Overwrites if 'out_file' already exists.
    """
    drive_service = get_drive_service()
    request = drive_service.files().export_media(
        fileId=spreadsheet_id,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
        if status:
            log_debug(f"Download {int(status.progress() * 100)}%.")

    out_file.write_bytes(fh.getvalue())
    log_debug(f"Downloaded spreadsheet to {out_file}")


# ---------------- XLSX PROCESSING (READ LOCAL FILE) ---------------- #
def get_all_sheet_names_local(xlsx_file: Path) -> List[str]:
    """Return a list of all sheet names from the local .xlsx file (cross-platform)."""
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    return wb.sheetnames


def fetch_movie_list_local(xlsx_file: Path, sheet_name: str) -> List[str]:
    """
    Read the first column (A) from 'sheet_name' in the local Excel file.
    Return a list of stripped movie names (non-empty).
    """
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    if sheet_name not in wb.sheetnames:
        log_debug(f"[ERROR] Sheet '{sheet_name}' not found in {xlsx_file.name}.")
        return []

    sheet = wb[sheet_name]
    movies = []
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).strip():
            movies.append(str(val).strip())
    return movies


# ---------------- JSON CREATION & POPULATION ---------------- #
def ensure_url_json_exists(sheet_name: str) -> Path:
    """
    Create an empty JSON file named <sheetName>Urls.json (with no spaces)
    inside TRAILERS_DIR if it doesn't exist. Return the file path.
    """
    safe_name = sanitize_filename(sheet_name).replace(" ", "")
    json_file = TRAILERS_DIR / f"{safe_name}Urls.json"
    if not json_file.exists():
        TRAILERS_DIR.mkdir(parents=True, exist_ok=True)
        json_file.write_text("{}", encoding="utf-8")
        log_debug(f"[INFO] Created file: {json_file}")
    return json_file


def populate_json_with_movies(sheet_name: str, movies: List[str]):
    """
    - Ensure <sheetName>Urls.json exists (no spaces).
    - Load its dictionary.
    - For each movie in 'movies', if not present, add with empty value "".
    - Write back the JSON in multiline format with no trailing comma.
    """
    json_file = ensure_url_json_exists(sheet_name)

    try:
        existing_data = json.loads(json_file.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        existing_data = {}

    for mv in movies:
        if mv not in existing_data:
            existing_data[mv] = ""

    # Write multiline JSON with no trailing comma
    lines = ["{"]
    keys = list(existing_data.keys())
    for i, key in enumerate(keys):
        comma = "," if i < len(keys) - 1 else ""
        lines.append(f'  "{key}": "{existing_data[key]}"{comma}')
    lines.append("}")

    json_file.write_text("\n".join(lines) + "\n", encoding="utf-8")
    log_debug(f"[INFO] Updated {json_file}")


# ---------------- LOCATE TRAILER (READ JSON OR SEARCH YT) ---------------- #
def locate_trailer(sheet_name: str, movie_title: str) -> (Optional[str], str, Optional[str]):
    """
    Look up a trailer URL in <sheetName>Urls.json, or fallback to YouTube search.
    Return (url, source, video_title).
    """
    safe_sheet = sanitize_filename(sheet_name).replace(" ", "")
    urls_file = TRAILERS_DIR / f"{safe_sheet}Urls.json"

    # 1. Check local JSON file
    if urls_file.exists():
        try:
            url_dict = json.loads(urls_file.read_text(encoding="utf-8"))
            normalized_dict = {normalize(k): v for k, v in url_dict.items()}
            key = normalize(movie_title)
            matched_url = (
                normalized_dict.get(key)
                or normalized_dict.get(fuzzy_search(key, list(normalized_dict.keys())) or '')
            )
            if matched_url and "youtube.com" in matched_url:
                return matched_url, "json", None
            elif matched_url:
                return matched_url, "json", None
        except json.JSONDecodeError as e:
            log_debug(f"[ERROR] JSON decoding failed: {e}")

    # 2. Fallback: YouTube search API
    api_result = youtube_api_search(movie_title + " official hd trailer")
    if api_result:
        api_url, yt_video_title = api_result
        return api_url, "youtube", yt_video_title

    # 3. No trailer found
    return None, "", None


# ---------------- YOUTUBE PLAYLIST CREATION ---------------- #
def get_youtube_service():
    """
    Create and return an authorized YouTube Data API client using OAuth (for user-level actions).
    Stores and reuses credentials in 'youtube_token.pickle'.
    """
    creds = None
    if YOUTUBE_TOKEN_FILE.exists():
        with open(YOUTUBE_TOKEN_FILE, "rb") as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(CLIENT_SECRET_FILE), YOUTUBE_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(YOUTUBE_TOKEN_FILE, "wb") as token:
            pickle.dump(creds, token)
    return build("youtube", "v3", credentials=creds)


def create_youtube_playlist(title: str, video_ids: List[str]) -> Optional[str]:
    """
    Create an unlisted YouTube playlist named 'title' and populate it with video_ids.
    Return the playlist URL or None on failure.
    """
    try:
        youtube = get_youtube_service()
        playlist = youtube.playlists().insert(
            part="snippet,status",
            body={
                "snippet": {
                    "title": title,
                    "description": "Auto-generated playlist by Movie Picker App",
                },
                "status": {"privacyStatus": "unlisted"}
            },
        ).execute()

        playlist_id = playlist["id"]
        for vid in video_ids:
            youtube.playlistItems().insert(
                part="snippet",
                body={
                    "snippet": {
                        "playlistId": playlist_id,
                        "resourceId": {
                            "kind": "youtube#video",
                            "videoId": vid
                        }
                    }
                },
            ).execute()

        return f"https://www.youtube.com/playlist?list={playlist_id}"
    except Exception as e:
        log_debug(f"[ERROR] YouTube playlist creation failed: {e}")
    return None


# ---------------- IMAGE HANDLING ---------------- #
def pick_random_movies(movies: List[str], count: int) -> List[str]:
    """Randomly sample 'count' distinct movies from the 'movies' list."""
    return random.SystemRandom().sample(movies, k=count)


def load_random_image(directory: Path, prefix: str, max_num: int):
    """
    Load a random image from 'directory' with a file name like 'prefix_1.png'
    up to 'prefix_{max_num}.png'. Return a PhotoImage or None if missing.
    """
    path = directory / f"{prefix}_{random.randint(1, max_num)}.png"
    if path.exists():
        return ImageTk.PhotoImage(Image.open(path))
    return None


def load_direction_image():
    """
    Load a random direction image (clockwise or counter_clockwise) from NUMBERS_DIR.
    """
    direction = random.choice(["clockwise", "counter_clockwise"])
    path = NUMBERS_DIR / f"{direction}.png"
    if path.exists():
        return ImageTk.PhotoImage(Image.open(path))
    return None


# ---------------- GUI SETUP ---------------- #
def center_window(window, width=800, height=800):
    """Center the 'window' on the screen at the specified width and height."""
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f"{width}x{height}+{x}+{y}")


def on_mousewheel(event):
    """
    Cross-platform mousewheel event. Windows/macOS typically use <MouseWheel>.
    Some Linux distros might need <Button-4>/<Button-5>.
    """
    scale = -1 * (event.delta // 120)  # typically 120 or 240 on Windows/mac
    middle_canvas.yview_scroll(scale, "units")


# ---------------- BUTTON LOGIC ---------------- #
def on_update_sheets():
    """
    1) Download the Google Sheet as 'ghib.xlsx'
    2) For each sheet in that file, read col A and write movie placeholders to JSON.
    """
    try:
        # Download the .xlsx
        download_spreadsheet_as_xlsx(SPREADSHEET_ID, GHIB_FILE)

        # For each sheet in the local XLSX, parse col A -> populate JSON
        sheet_names = get_all_sheet_names_local(GHIB_FILE)
        for sheet in sheet_names:
            movies = fetch_movie_list_local(GHIB_FILE, sheet)
            populate_json_with_movies(sheet, movies)

        messagebox.showinfo("Sheets Updated", f"Updated from Google Sheets!\nProcessed {len(sheet_names)} sheet(s).")
    except Exception as e:
        log_debug(f"[ERROR updating sheets]: {e}")
        messagebox.showerror("Error", "Failed to update from Google Sheets. Check logs.")


def on_start(event=None):
    """
    Main 'Start' button callback:
      - Ask for # of attendees and a sheet name.
      - Fuzzy-match the user sheet name to the local XLSX list (80% cutoff).
      - Load col A from that sheet, pick movies, build a YT playlist, etc.
    """
    # Validate # of attendees
    try:
        attendee_count = int(num_people_entry.get().strip())
        if attendee_count <= 0:
            raise ValueError
    except ValueError:
        return messagebox.showerror("Error", "Enter a positive integer for attendees.")

    # Get user sheet name, remove extra spaces, lowercase
    raw_sheet_input = sheet_name_entry.get().strip()
    if not raw_sheet_input:
        return messagebox.showerror("Error", "Provide a valid sheet name.")

    if not GHIB_FILE.exists():
        # If local XLSX is missing, auto-download
        try:
            download_spreadsheet_as_xlsx(SPREADSHEET_ID, GHIB_FILE)
        except Exception as e:
            log_debug(f"[ERROR auto-downloading xlsx]: {e}")
            return messagebox.showerror("Error", "No local XLSX found. Try 'Update Sheets' first.")

    # All actual sheet names
    actual_sheets = get_all_sheet_names_local(GHIB_FILE)

    # Create a map { normalized_name -> actual_sheet_name }
    # e.g. "mysheet" -> "My Sheet"
    # (strip spaces + lowercase)
    sheet_map = {}
    for s in actual_sheets:
        norm = re.sub(r"\s+", "", s.lower())
        sheet_map[norm] = s

    # 1) Try direct normal comparison
    user_normal = re.sub(r"\s+", "", raw_sheet_input.lower())
    chosen_sheet = sheet_map.get(user_normal)

    # 2) If not found, do fuzzy search on the keys
    if not chosen_sheet:
        possible_keys = list(sheet_map.keys())
        best_key = fuzzy_search(user_normal, possible_keys, cutoff=0.8)
        if best_key:
            chosen_sheet = sheet_map[best_key]

    if not chosen_sheet:
        return messagebox.showerror("Error", f"No sheet matched '{raw_sheet_input}' (80% cutoff).")

    # Now we have chosen_sheet as the actual sheet name
    movies = fetch_movie_list_local(GHIB_FILE, chosen_sheet)
    if not movies or attendee_count > len(movies):
        return messagebox.showerror("Error", f"Insufficient movie data in sheet '{chosen_sheet}'.")

    # Clear previous results in middle/right frames
    for frame in (middle_frame, right_frame):
        for widget in frame.winfo_children():
            widget.destroy()

    # Pick random movies
    selected_movies = pick_random_movies(movies, attendee_count + 1)
    playlist_video_ids = []

    # Build up the left vs. right columns for displaying results
    col_frame1 = tk.Frame(middle_frame, bg=BACKGROUND_COLOR)
    col_frame1.grid(row=0, column=0, sticky="nw")

    col_frame2 = tk.Frame(middle_frame, bg=BACKGROUND_COLOR)
    col_frame2.grid(row=0, column=1, sticky="nw", padx=20)

    current_column = col_frame1
    lines_in_column = 0
    max_lines_per_column = 20

    # Locate trailers and build the UI labels
    for movie in selected_movies:
        trailer, source, yt_video_title = locate_trailer(chosen_sheet, movie)
        display_text = movie
        color = "white"

        if source == "youtube" and yt_video_title:
            color = FALLBACK_COLOR
            display_text += f" (YT: {yt_video_title})"

        if trailer and "youtube.com/watch?v=" in trailer:
            # Extract the video ID
            video_id = trailer.split("v=")[-1].split("&")[0]
            playlist_video_ids.append(video_id)

            label = tk.Label(
                current_column, text=display_text, fg=color,
                cursor="hand2", wraplength=280, justify="left",
                bg=BACKGROUND_COLOR
            )
            label.pack(anchor="w", pady=2)
            label.bind("<Button-1>", lambda e, url=trailer: open_in_windows_default(url))
        else:
            label = tk.Label(
                current_column,
                text=f"{movie}: No trailer found",
                fg=ERROR_COLOR, wraplength=280, justify="left",
                bg=BACKGROUND_COLOR
            )
            label.pack(anchor="w", pady=2)

        root.update_idletasks()
        lines_in_column += 1
        # Switch to second column if we run out of vertical space
        if lines_in_column * (label.winfo_reqheight() + 4) >= middle_canvas.winfo_height() - 5 \
           and current_column == col_frame1:
            current_column = col_frame2
            lines_in_column = 0

    # If we have any trailer IDs, build a YouTube playlist
    if playlist_video_ids:
        title = f"Movie Night {datetime.date.today()}"
        playlist_url = create_youtube_playlist(title, playlist_video_ids)
        if playlist_url:
            open_in_windows_default(playlist_url)

        # Show random direction image
        direction_img = load_direction_image()
        if direction_img:
            dir_label = tk.Label(right_frame, image=direction_img, bg=BACKGROUND_COLOR)
            dir_label.image = direction_img  # keep a reference so it doesn't get GC'd
            dir_label.pack(pady=10)

        # Show random number image
        number_img = load_random_image(NUMBERS_DIR, "number", attendee_count)
        if number_img:
            num_label = tk.Label(right_frame, image=number_img, bg=BACKGROUND_COLOR)
            num_label.image = number_img
            num_label.pack(pady=10)
    else:
        messagebox.showinfo("No Trailers", "No valid trailers found.")


# ---------------- MAIN APP WITH SCROLL AND DARK MODE ---------------- #
root = tk.Tk()
root.title("Random Movie Picker")
root.configure(bg=BACKGROUND_COLOR)

left_frame = tk.Frame(root, padx=10, pady=10, bg=BACKGROUND_COLOR)
left_frame.pack(side="left", fill="y")

middle_container = tk.Frame(root, padx=10, pady=10, bg=BACKGROUND_COLOR)
middle_container.pack(side="left", fill="both", expand=True)

middle_canvas = tk.Canvas(middle_container, bg=BACKGROUND_COLOR, highlightthickness=0)
middle_canvas.pack(side="left", fill="both", expand=True)

scrollbar = tk.Scrollbar(middle_container, orient="vertical", command=middle_canvas.yview, bg=BACKGROUND_COLOR)
scrollbar.pack(side="right", fill="y")

middle_canvas.configure(yscrollcommand=scrollbar.set)
middle_canvas.bind('<Configure>', lambda e: middle_canvas.configure(scrollregion=middle_canvas.bbox("all")))

middle_frame = tk.Frame(middle_canvas, bg=BACKGROUND_COLOR)
middle_canvas.create_window((0, 0), window=middle_frame, anchor="nw")

# Cross-platform mousewheel binding
if sys.platform.startswith("win") or sys.platform == "darwin":
    # Windows or macOS usually support <MouseWheel>
    middle_canvas.bind_all("<MouseWheel>", on_mousewheel)
else:
    # Some Linux systems might need <Button-4>/<Button-5>
    middle_canvas.bind_all("<Button-4>", lambda e: middle_canvas.yview_scroll(-1, "units"))
    middle_canvas.bind_all("<Button-5>", lambda e: middle_canvas.yview_scroll(1, "units"))

right_frame = tk.Frame(root, padx=10, pady=10, bg=BACKGROUND_COLOR)
right_frame.pack(side="right", fill="y")

# Labels and Inputs
tk.Label(middle_frame, text="Movie Names", bg=BACKGROUND_COLOR, fg=FOREGROUND_COLOR)\
  .grid(row=0, column=0, columnspan=2, sticky="nw")

tk.Label(right_frame, text="Direction & Starting Pos", bg=BACKGROUND_COLOR, fg=FOREGROUND_COLOR).pack()

num_label = tk.Label(left_frame, text="Number of Attendees:", bg=BACKGROUND_COLOR, fg=FOREGROUND_COLOR)
num_label.pack(pady=5)

num_people_entry = tk.Entry(left_frame, bg=HIGHLIGHT_COLOR, fg=FOREGROUND_COLOR, insertbackground=FOREGROUND_COLOR)
num_people_entry.pack(pady=5)

sheet_label = tk.Label(left_frame, text="Sheet Name:", bg=BACKGROUND_COLOR, fg=FOREGROUND_COLOR)
sheet_label.pack(pady=5)

sheet_name_entry = tk.Entry(left_frame, bg=HIGHLIGHT_COLOR, fg=FOREGROUND_COLOR, insertbackground=FOREGROUND_COLOR)
sheet_name_entry.pack(pady=5)

start_button = tk.Button(
    left_frame,
    text="Start",
    command=on_start,
    bg=BUTTON_COLOR,
    fg=FOREGROUND_COLOR,
    activebackground=HIGHLIGHT_COLOR
)
start_button.pack(pady=5)

update_button = tk.Button(
    left_frame,
    text="Update Sheets",
    command=on_update_sheets,
    bg=BUTTON_COLOR,
    fg=FOREGROUND_COLOR,
    activebackground=HIGHLIGHT_COLOR
)
update_button.pack(pady=5)

root.bind('<Return>', on_start)
center_window(root)
root.mainloop()