from __future__ import annotations

import datetime
import importlib
import json
import os
import random
import re
import subprocess
import sys
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
import requests
from dotenv import load_dotenv
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from difflib import get_close_matches

import urllib.parse

#Gui 
try:
    from PySide6.QtCore import Qt, QUrl, Slot, QPropertyAnimation
    from PySide6.QtGui import QAction, QIcon, QColor, QPalette, QDesktopServices, QPixmap, QPainter, QFont, QColor   
    from PySide6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QListWidget, QListWidgetItem, QLabel, QStackedWidget, QPushButton,
        QLineEdit, QSplitter, QScrollArea, QTableWidget,
        QDialog, QCheckBox, QDialogButtonBox, QMessageBox, QGraphicsDropShadowEffect, QSizePolicy,
        QFrame, QGridLayout, QFrame,QProgressBar
    )
except ModuleNotFoundError as exc:
    sys.stderr.write(
        "PySide6 is not installed. GUI will not run.\n"
        "Install with:  pip install PySide6\n"
    )
    raise exc

#other python scripts needed in dir
import movie_repository


# ────────────────────────── Configuration ──────────────────────────
BASE_DIR = Path(__file__).resolve().parent

# file / folder paths
ENV_PATH            = BASE_DIR / "secret.env"
TRAILER_FOLDER      = BASE_DIR / "Video_Trailers"
NUMBERS_FOLDER      = BASE_DIR / "Numbers"
GHIBLI_SHEET_PATH   = BASE_DIR / "ghib.xlsx"
UNDER_REVIEW_PATH   = BASE_DIR / "underReviewURLs.json"
SERVICE_ACCOUNT_KEY = BASE_DIR / "service_secret.json"         # not used directly here
CLIENT_SECRET_PATH  = BASE_DIR / "client_secret.json"
USER_TOKEN_PATH     = BASE_DIR / "youtube_token.pickle"
LOG_PATH            = BASE_DIR / "trailer_debug.log"
AUTO_UPDATE_SCRIPT  = BASE_DIR / "autoUpdate.py"

# constants
ICON                = lambda name: QIcon(str(BASE_DIR / "icons" / f"{name}.svg"))
ACCENT_COLOR        = "#3b82f6"
YOUTUBE_SEARCH_URL  = "https://www.googleapis.com/youtube/v3/search"
DRIVE_SCOPES        = ["https://www.googleapis.com/auth/drive.readonly"]
YOUTUBE_SCOPES      = ["https://www.googleapis.com/auth/youtube"]

# env vars
load_dotenv(ENV_PATH)
SPREADSHEET_ID  = os.getenv("SPREADSHEET_ID")
YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY")
if not SPREADSHEET_ID or not YOUTUBE_API_KEY:
    raise EnvironmentError("Missing SPREADSHEET_ID or YOUTUBE_API_KEY in .env")

#Other files like probability (incase of changing name and such)
#-placeholder- 

# ────────────────────────── Logging helper ─────────────────────────
def log_debug(message: str) -> None:
    """
    Append a timestamped debug line to LOG_PATH.

    Uses Path.write_text(..., append=True) when available (3.12+),
    otherwise falls back to explicit 'a' mode open – so it never crashes.
    """
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.datetime.now().isoformat(timespec="seconds")
    log_line = f"[{timestamp}] {message}\n"

    try:
        # Python ≥ 3.12
        LOG_PATH.write_text(log_line, encoding="utf-8", append=True)  # type: ignore[arg-type]
    except TypeError:
        # Older Python – no 'append' kwarg
        with LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(log_line)

# ────────────────────────── Tiny utilities ─────────────────────────
def normalize(text: str) -> str:
    """Lower-case, strip, and drop non-alphanumerics (for fuzzy keys)."""
    return re.sub(r"[^a-z0-9]", "", text.lower().strip())


def fuzzy_match(target: str, candidates: List[str], cutoff: float = 0.8) -> Optional[str]:
    """Return best fuzzy match or None."""
    matches = get_close_matches(target, candidates, n=1, cutoff=cutoff)
    return matches[0] if matches else None
 
def make_number_pixmap(number: int,
                       size: int = 96,
                       fg_color: str = "#ffffff",
                       border_color: str = "#3b82f6") -> QPixmap:

    pixmap = QPixmap(size, size)
    pixmap.fill(Qt.transparent)

    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.Antialiasing)

    # draw border square
    pen = painter.pen()
    pen.setWidth(4)
    pen.setColor(QColor(border_color))
    painter.setPen(pen)
    painter.drawRoundedRect(2, 2, size - 4, size - 4, 10, 10)

    # draw number
    font = QFont("Arial", int(size * 0.55), QFont.Bold)
    painter.setFont(font)
    painter.setPen(QColor(fg_color))
    painter.drawText(pixmap.rect(), Qt.AlignCenter, str(number))
    painter.end()

    return pixmap

def resizeEvent(self, event):
    super().resizeEvent(event)          # default processing

    if not self.direction_label.pixmap():
        return  # nothing to scale yet

    # regenerate pixmaps at new size
    side = min(self.direction_label.width(), self.direction_label.height())
    arrow_pix = self._current_arrow_icon.pixmap(side, side)
    num_pix   = make_number_pixmap(self._current_number, size=side)

    self.direction_label.setPixmap(arrow_pix)
    self.number_label.setPixmap(num_pix)
    
# ───────────────────────── trailer-report helper ─────────────────────────
def report_trailer(movie_name: str, youtube_url: str | None) -> None:
    """
    Add/overwrite an entry in underReviewURLs.json and pop a confirmation.

    JSON is written pretty-printed (indented, one key per line) so it’s easy
    to read or edit by hand.
    """
    # 1) make sure the file exists
    if not UNDER_REVIEW_PATH.exists():
        UNDER_REVIEW_PATH.write_text("{}", encoding="utf-8")

    # 2) read existing contents (gracefully handle bad JSON)
    try:
        data: dict[str, str] = json.loads(UNDER_REVIEW_PATH.read_text("utf-8"))
    except json.JSONDecodeError:
        data = {}

    # 3) update / set the entry
    data[movie_name] = youtube_url or ""

    # 4) write it back – pretty JSON, Unix newlines, trailing newline
    UNDER_REVIEW_PATH.write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    # 5) log and notify
    log_debug(f"[REPORT] Marked “{movie_name}” → {youtube_url} as under review.")
    QMessageBox.information(
        None,
        "Trailer reported",
        f"Marked “{movie_name}” for manual review."
    )

# ────────────────────────── YouTube helpers ────────────────────────
def search_youtube_api(query: str) -> Optional[Tuple[str, str]]:
    """Return (url, video_title) for first short video result or None."""
    params = {
        "part": "snippet",
        "q": query,
        "key": YOUTUBE_API_KEY,
        "videoDuration": "short",
        "maxResults": 1,
        "type": "video",
    }
    try:
        response = requests.get(YOUTUBE_SEARCH_URL, params=params, timeout=10)
        data = response.json()
        if data.get("items"):
            first = data["items"][0]
            video_id = first["id"]["videoId"]
            title = first["snippet"]["title"]
            return f"https://www.youtube.com/watch?v={video_id}", title
    except Exception as exc:
        log_debug(f"YouTube search error: {exc}")
    return None


def locate_trailer(work_sheet: str, movie_title: str) -> Tuple[Optional[str], str, Optional[str]]:
    """
    Try local JSON first, then YouTube API.
    Returns (url, source, api_title) where source is 'json' or 'youtube'.
    """
    json_name = re.sub(r"\s+", "", work_sheet)
    urls_path = TRAILER_FOLDER / f"{json_name}Urls.json"

    if urls_path.exists():
        data = json.loads(urls_path.read_text(encoding="utf-8"))
        normalized = {normalize(k): v for k, v in data.items()}
        key = normalize(movie_title)
        url = normalized.get(key) or normalized.get(fuzzy_match(key, list(normalized)) or "")
        if url:
            return url, "json", None

    # fallback: YouTube
    api_result = search_youtube_api(f"{movie_title} official trailer")
    if api_result:
        url, api_title = api_result
        return url, "youtube", api_title

    return None, "", None


# ──────────────────────── OAuth & playlist helpers ─────────────────
def get_youtube_service():
    """Return cached or freshly-authenticated youtube client."""
    creds = None
    if USER_TOKEN_PATH.exists():
        creds = pickle.loads(USER_TOKEN_PATH.read_bytes())

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(CLIENT_SECRET_PATH), YOUTUBE_SCOPES)
            creds = flow.run_local_server(port=0)
        USER_TOKEN_PATH.write_bytes(pickle.dumps(creds))

    return build("youtube", "v3", credentials=creds)

def create_youtube_playlist(title: str, video_ids: List[str]) -> Optional[str]:
    """Create an unlisted playlist populated with `video_ids`; return playlist URL or None."""
    try:
        youtube = get_youtube_service()
        playlist = youtube.playlists().insert(
            part="snippet,status",
            body={
                "snippet": {
                    "title": title,
                    "description": "Auto-generated by Movie Night",
                },
                "status": {"privacyStatus": "unlisted"},
            },
        ).execute()

        playlist_id = playlist["id"]
        for vid in video_ids:
            youtube.playlistItems().insert(
                part="snippet",
                body={
                    "snippet": {
                        "playlistId": playlist_id,
                        "resourceId": {"kind": "youtube#video", "videoId": vid},
                    }
                },
            ).execute()

        return f"https://www.youtube.com/playlist?list={playlist_id}"
    except Exception as exc:
        log_debug(f"YouTube playlist creation failed: {exc}")
        return None

# ──────────────────────── Database Calls ─────────────────
DATABASE = movie_repository
def movie_probability(title: str) -> float:
    return DATABASE.get_prob(title)
def calculate_weighted_totals(titles):
    # to do will calculate combined weighted total, will call probability
    return DATABASE.get_calc_weighted_ratings(titles) 
def calculate_group_similarity(titles):
    return DATABASE.get_similarity(titles)  
# ────────────────────────── GUI widgets ───────────────────────────
class MovieCard(QFrame):
    """A single movie “card” with hover‐shadow animation."""
    def __init__(self,
                    title: str,
                    url: str,
                    probability: float,
                    parent: QWidget | None = None):
        super().__init__(parent)
        self.setObjectName("MovieCardItem")
        self.setFrameShape(QFrame.StyledPanel)
        self.setFrameShadow(QFrame.Raised)

        # layout
        lay = QVBoxLayout(self)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(6)
        
         # Title (clickable link)
        title_lbl = QLabel(f'<a href="{url}">{title}</a>')
        title_lbl.setTextFormat(Qt.RichText)
        title_lbl.setTextInteractionFlags(Qt.TextBrowserInteraction)
        title_lbl.setOpenExternalLinks(True)
        title_lbl.setWordWrap(True)
        title_lbl.setProperty("class", "MovieTitle")
        lay.addWidget(title_lbl)

        # Probability pill with background colour and padding
        prob_str = f"{probability:.2f}"
        prob_lbl = QLabel(prob_str, alignment=Qt.AlignCenter)

        # pick background colour
        if   probability >= 0.7:
            pill_bg = "#2ecc71"    # green
            text_color = "#000"
        elif probability >= 0.4:
            pill_bg = "#f1c40f"    # yellow
            text_color = "#000"
        else:
            pill_bg = ACCENT_COLOR # your blue
            text_color = "#fff"

        # inline style for the “pill”
        prob_lbl.setStyleSheet(f"""
            background: {pill_bg};
            color: {text_color};
            border-radius: 6px;
            padding: 2px 8px;
            min-width: 32px;
        """)

        lay.addWidget(prob_lbl, alignment=Qt.AlignHCenter)


        lay.addStretch()

        # drop-shadow effect
        self._shadow = QGraphicsDropShadowEffect(self)
        self._shadow.setBlurRadius(4)
        self._shadow.setColor(QColor(0, 0, 0, 100))
        self._shadow.setOffset(0, 0)
        self.setGraphicsEffect(self._shadow)
        
    def enterEvent(self, event):
        super().enterEvent(event)
        anim = QPropertyAnimation(self._shadow, b"blurRadius", self)
        anim.setDuration(200)
        anim.setEndValue(16)
        anim.start(QPropertyAnimation.DeleteWhenStopped)

    def leaveEvent(self, event):
        super().leaveEvent(event)
        anim = QPropertyAnimation(self._shadow, b"blurRadius", self)
        anim.setDuration(200)
        anim.setEndValue(4)
        anim.start(QPropertyAnimation.DeleteWhenStopped)
        
class ReportDialog(QDialog):
    """Checkbox list of movies to mark as ‘bad trailer’."""

    def __init__(self, parent: QWidget, movies: List[str]) -> None:
        super().__init__(parent)
        self.setWindowTitle("Report trailers")

        layout = QVBoxLayout(self)
        self._boxes: List[QCheckBox] = [QCheckBox(title) for title in movies]
        for box in self._boxes:
            layout.addWidget(box)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_movies(self) -> List[str]:
        return [box.text() for box in self._boxes if box.isChecked()]


class PickerPage(QWidget):
    """Main picker UI (left controls, center movie list, right stats)."""

    DIRECTIONS = ["Clockwise", "Counter-Clockwise"]

    def __init__(self, main_window: "MainWindow") -> None:
        super().__init__()

        self.main_window = main_window
        self._build_layout()
        self._connect_signals()

        # show direction/number only after first Generate Movies
        self.direction_label.hide()
        self.number_label.hide()

    # ───────────── layout helpers ─────────────
    def _build_layout(self) -> None:
        outer = QHBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(20)        

        # left: controls
        control_layout = QVBoxLayout()
        self.attendee_input = QLineEdit(placeholderText="# attendees")
        self.sheet_input = QLineEdit(placeholderText="Sheet name")

        self.generate_button = QPushButton("Generate Movies")
        self.update_urls_button = QPushButton("Update URLs")
        
        for b in (self.generate_button, self.update_urls_button):
            b.setAutoDefault(False)
            b.setFlat(True)     # flat = auto-raise style in Fusion


   # ─── Build controls + stats card ────────────────────
        for widget in (
            self.attendee_input,
            self.sheet_input,
            self.generate_button,
            self.update_urls_button,
        ):
            control_layout.addWidget(widget)

        # ─── Stats card sits immediately after the Update button ─
        stats_card = QFrame()
        stats_card.setObjectName("StatsCard")
        stats_layout = QVBoxLayout(stats_card)
        stats_layout.setContentsMargins(8, 8, 8, 8)
        stats_layout.setSpacing(4)

        header = QLabel("Group Metrics", alignment=Qt.AlignCenter)
        header.setProperty("class", "StatsHeader")

        self.group_sim_label = QLabel("Similarity: —", alignment=Qt.AlignCenter)
        self.group_sim_label.setProperty("class", "StatsValue")

        self.similarity_bar = QProgressBar()
        self.similarity_bar.setRange(0, 100)
        self.similarity_bar.setTextVisible(False)

        self.group_weighted_score = QLabel("Weighted Score: —", alignment=Qt.AlignCenter)
        self.group_weighted_score.setProperty("class", "StatsValue")

        stats_layout.addWidget(header)
        stats_layout.addWidget(self.group_sim_label)
        stats_layout.addWidget(self.similarity_bar)
        stats_layout.addWidget(self.group_weighted_score)

        control_layout.addWidget(stats_card)
        control_layout.addStretch()          # push stats (and controls) upward
        outer.addLayout(control_layout)
        
        # centre: scrollable movie grid
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)

        # container for cards
        self.movie_grid_frame = QWidget()
        self.movie_grid_frame.setObjectName("MovieCardContainer")
        self.movie_grid_layout = QGridLayout(self.movie_grid_frame)
        self.movie_grid_layout.setContentsMargins(12, 12, 12, 12)
        self.movie_grid_layout.setSpacing(12)

        self.scroll_area.setWidget(self.movie_grid_frame)
        outer.addWidget(self.scroll_area, 2)


        # right: stats & direction
        right_layout = QVBoxLayout()

        self.stats_label = QLabel(alignment=Qt.AlignCenter)
        right_layout.addWidget(self.stats_label)

        # direction + number tiles
        self.direction_label = QLabel("", alignment=Qt.AlignCenter, objectName="DirectionTile")
        self.number_label = QLabel("", alignment=Qt.AlignCenter, objectName="NumberTile")   
        for lbl in (self.direction_label, self.number_label):
            lbl.setMinimumSize(80, 80)                         # keeps them usable when tiny
            lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            lbl.setScaledContents(True)   
        self.direction_label.setStyleSheet("border: 2px solid #555; border-radius: 8px;")

        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(12)
        shadow.setColor(QColor(0, 0, 0, 160))   # semi-transparent black
        shadow.setOffset(0, 2)
        self.number_label.setGraphicsEffect(shadow)
        
        right_layout.addWidget(self.direction_label, alignment=Qt.AlignHCenter)
        right_layout.addWidget(self.number_label, alignment=Qt.AlignHCenter)

        self.report_button = QPushButton("Report Trailers")
        self.report_button.setEnabled(False)
        right_layout.addWidget(self.report_button)
        right_layout.addStretch()

        outer.addLayout(right_layout)

    def _connect_signals(self) -> None:
        self.generate_button.clicked.connect(self.main_window.generate_movies)
        self.update_urls_button.clicked.connect(self.main_window.update_urls)
        self.report_button.clicked.connect(self._open_report_dialog)

        # pressing Enter in either LineEdit triggers generate
        self.attendee_input.returnPressed.connect(self.main_window.generate_movies)
        self.sheet_input.returnPressed.connect(self.main_window.generate_movies)

    # ───────────── public API ─────────────
    def display_movies(self, movies: List[str], trailer_lookup: dict[str, str]) -> None:
        # clear old cards
        while self.movie_grid_layout.count():
            item = self.movie_grid_layout.takeAt(0)
            if widget := item.widget():
                widget.deleteLater()
        self.current_movies = movies
        self.current_lookup = trailer_lookup
        
        # determine card width (including spacing)
        card_width = 200                
        margins    = self.movie_grid_layout.contentsMargins()
        spacing    = self.movie_grid_layout.horizontalSpacing()

        # fetch the inner viewport width of the scroll area
        available_width = self.scroll_area.viewport().width() \
                        - margins.left() - margins.right()
                        
        # compute number of columns, at least 1
        cols = max(1, (available_width + spacing) // (card_width + spacing))

        # add a MovieCard for each movie
        for idx, title in enumerate(movies):
            url        = trailer_lookup.get(title, "")
            prob       = movie_probability(title)
            card       = MovieCard(title, url, prob, self)
            row, col   = divmod(idx, cols)
            self.movie_grid_layout.addWidget(card, row, col)

        # random direction + number
        # inside PickerPage.display_movies or wherever you pick direction/number
        direction = random.choice(self.DIRECTIONS)                    # "Clockwise" / "Counter-Clockwise"
        icon_map = {
            "Clockwise": "arrow-clockwise",
            "Counter-Clockwise": "arrow-counterclockwise",
        }
        icon_name = icon_map[direction]                               # e.g. "arrow-right-circle"

        number = random.randint(1, (int(self.attendee_input.text())))
        label_w   = self.direction_label.width()
        label_h   = self.direction_label.height()
        side      = min(label_w, label_h, 124) 
        # build pixmaps
        arrow_icon    = ICON(icon_name)
        arrow_pix     = arrow_icon.pixmap(side,side)
        number_pixmap = make_number_pixmap(
                        number,
                        size = side,
                        fg_color="#ffffff",                     
                        ) 

        # push to labels
        self.direction_label.setPixmap(arrow_pix)
        self.number_label.setPixmap(number_pixmap)

        # optional: let Qt scale when label is resized
        self.direction_label.setScaledContents(True)
        self.number_label.setScaledContents(True)

        self.direction_label.show()
        self.number_label.show()
    
        self._current_icon_name = arrow_icon   # <<< store icon
        self._current_number     = number                                 # <<< store value

        #similaraties and group metascore
        similarity = calculate_group_similarity(movies)
        weighted_scores = calculate_weighted_totals(movies)  
        self.group_sim_label.setText(f"Similarity: {similarity*100:.1f}%")
        self.group_weighted_score.setText(f"Total_Rating: {weighted_scores*100:.1f}" )
        #update bar and movies
        sim_pct = int(similarity * 100)
        self.group_sim_label.setText(f"Similarity: {sim_pct}%")
        self.similarity_bar.setValue(sim_pct)
        self.group_weighted_score.setText(f"Weighted Score: {weighted_scores*100:.1f}")
        
    # ───────────── internal slots ─────────────
    @Slot()
    def _open_report_dialog(self) -> None:
        dialog = ReportDialog(self, self.current_movies)
        if dialog.exec() == QDialog.Accepted:
            for title in dialog.selected_movies():
                report_trailer(title, self.current_lookup.get(title, ""))
            QMessageBox.information(self, "Reported", "Thanks for the feedback!")


class StatsPage(QWidget):
    """Placeholder stats page."""
    def __init__(self) -> None:
        super().__init__()
        self.table = QTableWidget(0, 3)  # TODO: implement real stats


# ────────────────────────── Main Window ───────────────────────────
class MainWindow(QMainWindow):

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Movie Night")
        self.resize(960, 600)

        self.picker_page = PickerPage(self)
        self.stats_page = StatsPage()

        # navigation list
        self.nav_list = QListWidget()
        self.nav_list.setFixedWidth(170)
        self.nav_list.addItem(QListWidgetItem(ICON("grid"), "RNJesus"))
        row = self.nav_list.count()                # index of the item you just appended
        item = self.nav_list.item(row - 1)         # last added item
        item.setTextAlignment(Qt.AlignHCenter)
        self.nav_list.addItem(QListWidgetItem(ICON("bar-chart-line"), "  Movie Stats"))
        row = self.nav_list.count()                # index of the item you just appended
        item = self.nav_list.item(row - 1)         # last added item
        item.setTextAlignment(Qt.AlignHCenter)


        # stacked pages
        self.pages = QStackedWidget()
        self.pages.addWidget(self.picker_page)
        self.pages.addWidget(self.stats_page)

        self.nav_list.currentRowChanged.connect(self.pages.setCurrentIndex)

        splitter = QSplitter()
        splitter.addWidget(self.nav_list)
        splitter.addWidget(self.pages)
        splitter.setStretchFactor(1, 1)
        self.setCentralWidget(splitter)
        splitter.setHandleWidth(1)  

        # toolbar
        toolbar = self.addToolBar("Main")
        reroll_action = QAction(ICON("refresh"), "Re-roll", self)
        reroll_action.setShortcut("Ctrl+R")
        reroll_action.triggered.connect(self.generate_movies)
        toolbar.addAction(reroll_action)

    # ─────────────────── public slots ───────────────────
    @Slot()
    def update_urls(self) -> None:
        subprocess.run(["python", AUTO_UPDATE_SCRIPT], check=False)

    @Slot()
    def generate_movies(self) -> None:
        """Validate input, pick movies, build playlist, update UI."""
        try:
            attendee_count = int(self.picker_page.attendee_input.text().strip())
            if attendee_count <= 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "Error", "Enter a positive attendee count.")
            return

        sheet_name_raw = self.picker_page.sheet_input.text().strip()
        if not sheet_name_raw:
            QMessageBox.warning(self, "Error", "Sheet name?")
            return
        if not GHIBLI_SHEET_PATH.exists():
            QMessageBox.warning(self, "Error", "Run Update URLs first.")
            return

        workbook = openpyxl.load_workbook(GHIBLI_SHEET_PATH, read_only=True)
        sheet_map = {normalize(name): name for name in workbook.sheetnames}
        chosen_sheet = (
            sheet_map.get(normalize(sheet_name_raw))
            or sheet_map.get(fuzzy_match(normalize(sheet_name_raw), list(sheet_map)))
        )
        if not chosen_sheet:
            QMessageBox.warning(self, "Error", "Sheet not found.")
            return

        movie_titles = [
            row[0] for row in workbook[chosen_sheet].iter_rows(min_row=1, max_col=1, values_only=True)
            if row[0]
        ]
        if attendee_count + 1 > len(movie_titles):
            QMessageBox.warning(self, "Error", "Not enough movies.")
            return

        chosen_movies = random.sample(movie_titles, attendee_count + 1)
        trailer_lookup = {
            title: locate_trailer(chosen_sheet, title)[0] for title in chosen_movies
        }

        # optional: build YouTube playlist from found trailers
        video_ids = [
            url.split("v=")[-1].split("&")[0]
            for url in trailer_lookup.values()
            if url and "watch?v=" in url
        ]
        if video_ids:
            ids_csv = ",".join(video_ids)
            title   = urllib.parse.quote_plus(f"Movie Night {datetime.date.today()}")
            playlist_link = (
                f"https://www.youtube.com/watch_videos"
                f"?video_ids={ids_csv}"
                f"&title={title}"
                f"&feature=share"
            )
            QDesktopServices.openUrl(QUrl(playlist_link))
        # push to UI
        self.picker_page.display_movies(chosen_movies, trailer_lookup)


# ────────────────────────── Dark theme ────────────────────────────
def apply_dark_palette(app: QApplication) -> None:
    """Dark Fusion palette with brand accent."""
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor("#202124"))
    palette.setColor(QPalette.WindowText, Qt.white)
    palette.setColor(QPalette.Base, QColor("#2b2c2e"))
    palette.setColor(QPalette.AlternateBase, QColor("#323336"))
    palette.setColor(QPalette.Button, QColor("#2d2e30"))
    palette.setColor(QPalette.ButtonText, Qt.white)
    palette.setColor(QPalette.Text, Qt.white)
    palette.setColor(QPalette.Link, QColor(ACCENT_COLOR))
    palette.setColor(QPalette.Highlight, QColor(ACCENT_COLOR))
    palette.setColor(QPalette.HighlightedText, Qt.white)
    app.setStyle("Fusion")
    app.setPalette(palette)


# ────────────────────────── Entry point ───────────────────────────
def main() -> None:
    app = QApplication([])
    apply_dark_palette(app)
    app.setStyleSheet("""
    /* ---- base typography ---- */
    QWidget            { font-family:"Inter","Roboto","Arial"; font-size:12pt; }
    QLabel.MovieTitle  { font-weight:600; font-size:13pt; }

    /* ---- movie-row hover ---- */
    QWidget#MovieRow:hover        { background:#303030; }
    QWidget#MovieRow:hover a      { text-decoration:underline; }

    /* ---- rounded tile gradient ---- */
    QLabel#DirectionTile, QLabel#NumberTile {
        border               :2px solid transparent;
        background           :qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #555, stop:1 #333);
        border-radius        :10px;
    }
    
    /* base styling for each MovieCard */
    QFrame#MovieCardItem {
        background: #2b2c2e;
        border-radius: 8px;}
    /*Hover Movie Card*/    
    QFrame#MovieCardItem:hover {background-color: #303030;}

    /* ---- flat buttons (auto-raise look) ---- */
    QPushButton          { background:transparent; border:1px solid #555; padding:3px 10px; }
    QPushButton:hover    { background:#404040; }

    /* ---- slim vertical scrollbar ---- */
    QScrollBar:vertical            { width:8px; background:transparent; }
    QScrollBar::handle:vertical    { background:#555; border-radius:4px; }
    QScrollBar::add-line, QScrollBar::sub-line { height:0; }
    QScrollBar::add-page, QScrollBar::sub-page { background:transparent; }

    /* ---- splitter handle invisible ---- */
    QSplitter::handle    { background:transparent; }
    
    QWidget#MovieCardContainer {
        background: #272727;
        border-radius: 12px;
    }

    /* individual card hover cursor */
    QFrame#MovieCardItem { cursor: pointer;}
    
    /* ─── stats card styling ──────*/
    QFrame#StatsCard {
        background: #272727;
        border: 1px solid #555;
        border-radius: 8px;}
    
    /* header */
    QLabel.StatsHeader {
        color: #eee;
        font-weight: 600;
        font-size: 11pt;}
    
    /* values */
    QLabel.StatsValue {
        color: #3b82f6;     
        font-weight: 700;
        font-size: 12pt;
        padding: 2px 0;}
    
    QProgressBar {
        border: 1px solid #444; border-radius: 4px;
        background: #202124;
        height: 10px;
        margin: 4px 0;}
  
    QProgressBar::chunk {
        background: #3b82f6;
        border-radius: 4px;}
  
    """
    )
    
    MainWindow().show()
    app.exec()


if __name__ == "__main__":
    main()
