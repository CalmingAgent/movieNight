import os
import sys
import json
import re
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv
import datetime

# ----------------- CONFIGURATION ----------------- #
BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / "secret.env"

# Load environment variables from secret.env
load_dotenv(ENV_PATH)

TMDB_API_KEY = os.getenv("TMDB_API_KEY")
YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY")

if not TMDB_API_KEY:
    raise EnvironmentError("Missing 'TMDB_API_KEY' in secret.env or environment variables.")
if not YOUTUBE_API_KEY:
    raise EnvironmentError("Missing 'YOUTUBE_API_KEY' in secret.env or environment variables.")

# Local Excel file
GHIB_FILE = BASE_DIR / "ghib.xlsx"

# Where JSON trailer files are stored
TRAILERS_DIR = BASE_DIR / "Video_Trailers"
TRAILERS_DIR.mkdir(parents=True, exist_ok=True)

LOG_FILE = BASE_DIR / "trailer_debug.log"

# ---------------- GLOBAL FLAG FOR YOUTUBE QUOTA ---------------- #
YOUTUBE_MAXED_OUT = False  # if True, skip further YT lookups

# --------------- LOGGING & UTILS --------------- #
def log_debug(message: str) -> None:
    """Log debug messages to a file for troubleshooting (cross-platform)."""
    timestamp = datetime.datetime.now().isoformat(timespec='seconds')
    with open(LOG_FILE, "a", encoding="utf-8") as log_file:
        log_file.write(f"[{timestamp}] {message}\n")

def sanitize_filename(name: str) -> str:
    """Remove characters invalid for filenames, strip leading/trailing spaces."""
    return re.sub(r'[<>:"/\\|?*]', '', name.strip())

def normalize_title(title: str) -> str:
    """
    Convert to lowercase, strip leading/trailing whitespace,
    and remove all non-alphanumeric characters.
    """
    return re.sub(r'[^a-z0-9]', '', title.lower().strip())

# -------------- PROGRESS BAR -------------- #
def print_progress_bar(iteration: int, total: int, prefix: str="", suffix: str="", length: int=40):
    """
    Print a progress bar to the console.
    - iteration: current iteration (int)
    - total: total iterations (int)
    - prefix: string to prefix progress bar
    - suffix: string to suffix progress bar
    - length: character length of the bar
    """
    if total <= 0:
        return
    fraction = iteration / float(total)
    filled_length = int(length * fraction)
    bar = "█" * filled_length + "-" * (length - filled_length)
    percent = round(100 * fraction, 1)
    print(f"\r{prefix} |{bar}| {percent}% {suffix}", end="\r")
    if iteration >= total:
        print()

# -------------- XLSX READING -------------- #
def get_all_sheet_names_local(xlsx_file: Path):
    """Return a list of all sheet names from the local XLSX file."""
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    return wb.sheetnames

def fetch_movie_list_local(xlsx_file: Path, sheet_name: str):
    """
    Read the first column (A) from 'sheet_name' in the local Excel file.
    Return a list of *stripped* movie names (non-empty).
    """
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    if sheet_name not in wb.sheetnames:
        log_debug(f"[ERROR] Sheet '{sheet_name}' not found in {xlsx_file.name}.")
        return []

    sheet = wb[sheet_name]
    movies = []
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).strip():
            movies.append(str(val).strip())  # strip leading/trailing
    return movies

# ---------------- TMDB + YOUTUBE WITH QUOTA CHECKS ---------------- #
def tmdb_find_trailer(movie_title: str):
    """
    1) Fetch the first 2 pages of TMDB /search/movie results for 'movie_title'.
    2) Combine the results, log how many total results.
    3) Filter for EXACT normalized matches of the "title".
    4) If no matches or >=3 matches, log and return None (fallback to YT).
    5) Otherwise, use the first matched's ID -> fetch /videos, return the first YT link.
    If TMDB returns 429, we exit. 
    Logs each step for debugging.
    """
    try:
        norm_query = normalize_title(movie_title)
        all_results = []
        # We'll do 2 pages max
        for page in [1, 2]:
            search_url = "https://api.themoviedb.org/3/search/movie"
            params = {
                "api_key": TMDB_API_KEY,
                "query": movie_title,
                "page": page
            }
            log_debug(f"[TMDB] Searching page={page} for '{movie_title}'")
            resp = requests.get(search_url, params=params, timeout=10)

            if resp.status_code == 429:
                log_debug("[TMDB] 429 Too Many Requests - daily limit reached. Exiting.")
                print("TMDB daily quota limit reached. Exiting.")
                sys.exit(1)

            data = resp.json()
            page_results = data.get("results", [])
            if not page_results:
                break  # no more results, so stop
            all_results.extend(page_results)

            # If total_pages or results are done, we can break early
            total_pages = data.get("total_pages", 1)
            if page >= total_pages:
                break

        log_debug(f"[TMDB] Found {len(all_results)} total results (pages 1-2) for '{movie_title}'")

        if not all_results:
            log_debug(f"[TMDB] No results at all for '{movie_title}'; fallback to YT.")
            return None

        # Filter to exact normalized matches
        matched = []
        for r in all_results:
            rtitle = r.get("title", "")
            if normalize_title(rtitle) == norm_query:
                matched.append(r)

        log_debug(f"[TMDB] Found {len(matched)} EXACT matches for '{movie_title}' (normalize='{norm_query}')")

        # If no or >=3 exact matches, fallback to YouTube
        if len(matched) == 0:
            log_debug(f"[TMDB] Zero exact matches => fallback to YT for '{movie_title}'")
            return None
        if len(matched) >= 3:
            log_debug(f"[TMDB] {len(matched)} exact matches (>=3) => fallback to YT for '{movie_title}'")
            return None

        # Use the first matched
        chosen = matched[0]
        movie_id = chosen["id"]
        log_debug(f"[TMDB] Using matched ID={movie_id} for '{chosen.get('title')}'")

        # Now fetch videos
        videos_url = f"https://api.themoviedb.org/3/movie/{movie_id}/videos"
        vid_params = {"api_key": TMDB_API_KEY}
        vids_resp = requests.get(videos_url, params=vid_params, timeout=10)

        if vids_resp.status_code == 429:
            log_debug("[TMDB] 429 Too Many Requests while fetching /videos - daily limit reached. Exiting.")
            print("TMDB daily quota limit reached. Exiting.")
            sys.exit(1)

        vids_data = vids_resp.json().get("results", [])
        # Return the first trailer link
        for vid in vids_data:
            if vid.get("site") == "YouTube" and vid.get("type") in ("Trailer", "Teaser"):
                key = vid.get("key")
                if key:
                    log_debug(f"[TMDB] Found YT trailer key={key} for '{movie_title}' using matched ID={movie_id}")
                    return f"https://www.youtube.com/watch?v={key}"

        log_debug(f"[TMDB] No YT trailer found in /videos for '{movie_title}' => fallback to YT search.")
        return None
    except Exception as e:
        log_debug(f"[ERROR] TMDB search error for '{movie_title}': {e}")
        return None

def youtube_api_search(query: str):
    """
    Use the YouTube Data API (v3) to find a single short video by 'query'.
    If daily limit/quota is exceeded (403), we set YOUTUBE_MAXED_OUT = True
    and return None (skip writing, but do NOT exit).
    Logs each step for debugging.
    """
    global YOUTUBE_MAXED_OUT

    if YOUTUBE_MAXED_OUT:
        # Already maxed, skip calling
        log_debug(f"[YouTube] Skipping search for '{query}' because YOUTUBE_MAXED_OUT is True")
        return None

    try:
        log_debug(f"[YouTube] Searching for '{query}'")
        search_url = "https://www.googleapis.com/youtube/v3/search"
        params = {
            "part": "snippet",
            "q": query,
            "key": YOUTUBE_API_KEY,
            "videoDuration": "short",
            "maxResults": 1,
            "type": "video"
        }
        resp = requests.get(search_url, params=params, timeout=10)

        if resp.status_code == 403:
            error_data = resp.json().get("error", {})
            reason_list = error_data.get("errors", [])
            for er in reason_list:
                reason = er.get("reason", "")
                if reason in ("quotaExceeded", "dailyLimitExceeded"):
                    log_debug("[YouTube] 403 daily quota limit reached - setting YOUTUBE_MAXED_OUT=True, skipping.")
                    print("YouTube daily quota limit reached. Skipping further YT lookups.")
                    YOUTUBE_MAXED_OUT = True
                    return None
            log_debug(f"[YouTube] 403 error but not quota exceeded. reason_list={reason_list}")
            return None

        data = resp.json()
        items = data.get("items", [])
        if items:
            video_id = items[0]["id"]["videoId"]
            video_title = items[0]["snippet"]["title"]
            log_debug(f"[YouTube] Found video_id={video_id} for '{query}' => {video_title}")
            return f"https://www.youtube.com/watch?v={video_id}", video_title
        else:
            log_debug(f"[YouTube] No items found for '{query}'")
            return None
    except Exception as e:
        log_debug(f"[ERROR] YouTube API error for '{query}': {e}")
        return None

def find_trailer_fallback_cache(movie_title: str, master_cache: dict):
    """
    1) Normalize and check master_cache first.
    2) If not found, try tmdb_find_trailer.
    3) If that returns None, fallback to YouTube (unless YT is maxed).
    4) Return the found URL or None, store in cache if found.
    Logs which API ends up providing the result.
    """
    norm = normalize_title(movie_title)
    if norm in master_cache and master_cache[norm]:
        log_debug(f"[CACHE] Using cached URL for '{movie_title}' => {master_cache[norm]}")
        return master_cache[norm]

    # Try TMDB
    tmdb_url = tmdb_find_trailer(movie_title)
    if tmdb_url:
        master_cache[norm] = tmdb_url
        log_debug(f"[CACHE] Storing TMDB result for '{movie_title}' => {tmdb_url}")
        return tmdb_url

    # Fallback to YouTube
    yt_query = movie_title + " official hd trailer"
    yt_result = youtube_api_search(yt_query)
    if yt_result:
        video_url, _ = yt_result
        master_cache[norm] = video_url
        log_debug(f"[CACHE] Storing YouTube result for '{movie_title}' => {video_url}")
        return video_url

    # Neither worked
    log_debug(f"[FALLBACK] No trailer found from TMDB or YouTube for '{movie_title}'")
    return None

# -------------- MASTER CACHE -------------- #
def build_master_cache_from_all_json() -> dict:
    """
    Reads all JSON files in TRAILERS_DIR, collecting known URLs into:
      { normalized_movie_title: "YouTube URL" }
    for quick reuse across duplicates.
    """
    master_cache = {}
    for json_file in TRAILERS_DIR.glob("*.json"):
        try:
            text = json_file.read_text(encoding="utf-8")
            data = json.loads(text)
        except json.JSONDecodeError:
            data = {}

        for title, url in data.items():
            if url:
                # Normalize the key for duplicates
                norm = normalize_title(title)
                if norm not in master_cache:
                    master_cache[norm] = url

    return master_cache

# -------------- JSON HELPERS -------------- #
def load_json_dict(json_path: Path):
    """Load a JSON file into a dictionary (or return empty dict if parse error)."""
    try:
        return json.loads(json_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}

def write_json_dict(json_path: Path, data: dict):
    """Write `data` to `json_path` in multiline JSON format (no trailing commas)."""
    lines = ["{"]
    keys = list(data.keys())
    for i, key in enumerate(keys):
        comma = "," if i < len(keys) - 1 else ""
        val = data[key] or ""
        val_escaped = val.replace('"', '\\"')
        lines.append(f'  "{key}": "{val_escaped}"{comma}')
    lines.append("}")

    json_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

def ensure_url_json_exists(sheet_name: str) -> Path:
    """
    Create (if needed) <sheetName>Urls.json in TRAILERS_DIR, sanitized for filename.
    """
    safe_sheet_name = sanitize_filename(sheet_name).replace(" ", "")
    json_file = TRAILERS_DIR / f"{safe_sheet_name}Urls.json"
    if not json_file.exists():
        json_file.write_text("{}", encoding="utf-8")
        log_debug(f"[INFO] Created file: {json_file}")
    return json_file

# -------------- FILL MISSING URLS -------------- #
def fill_missing_urls_in_json_with_cache(json_file: Path, movies: list, master_cache: dict):
    """
    - Load the JSON from `json_file`.
    - Ensure placeholders for all `movies` from the sheet.
    - For any movie with empty URL, check master cache -> TMDB -> (optionally) YouTube.
    - If YT is maxed, we skip further YT calls for the rest of the run.
    - Write the JSON after updating (only for new links found).
    """
    data = load_json_dict(json_file)

    # Ensure placeholders for all movies
    for m in movies:
        if m not in data:
            data[m] = ""

    # Gather missing
    missing_entries = [m for m in movies if not data[m]]
    total_missing = len(missing_entries)
    if total_missing == 0:
        log_debug(f"[INFO] No missing URLs for {json_file.name}.")
        print(f"No missing URLs in {json_file.name} — skipping searches.")
        return

    print(f"\nFilling missing URLs in {json_file.name}: {total_missing} to search.")
    updated = False

    for i, movie_title in enumerate(missing_entries, start=1):
        print_progress_bar(i - 1, total_missing, prefix="Progress", suffix="Complete")

        # Attempt to find or reuse a trailer
        trailer_url = find_trailer_fallback_cache(movie_title, master_cache)
        if trailer_url:
            data[movie_title] = trailer_url
            updated = True
            log_debug(f"[INFO] Filled '{movie_title}' with URL: {trailer_url}")
        else:
            log_debug(f"[INFO] No trailer found for '{movie_title}' (TMDB/YT)")

    print_progress_bar(total_missing, total_missing, prefix="Progress", suffix="Complete")

    if updated:
        # Write the JSON, as we did find some new trailers
        write_json_dict(json_file, data)
        log_debug(f"[INFO] Updated JSON: {json_file}")
        print(f"Completed filling missing trailers in {json_file.name}")
    else:
        print(f"No valid trailer links found for any missing titles in {json_file.name}")

# -------------- MAIN PROCESS -------------- #
def fill_missing_urls_for_all_sheets_with_duplicates_tmdb():
    """
    1) Build a master cache from all JSON in TRAILERS_DIR to handle duplicates globally (normalized).
    2) For each sheet in 'ghib.xlsx', load/create a JSON file.
    3) Fill missing URLs using the cache -> TMDB -> (optionally) YouTube approach.
       - If daily limit is reached on TMDB, we exit (unchanged from the original).
       - If daily limit is reached on YT, set YOUTUBE_MAXED_OUT=True and skip further YT calls.
    """
    if not GHIB_FILE.exists():
        print(f"Error: local XLSX '{GHIB_FILE}' not found. Exiting.")
        sys.exit(1)

    master_cache = build_master_cache_from_all_json()

    sheet_names = get_all_sheet_names_local(GHIB_FILE)
    if not sheet_names:
        print("No sheets found in the Excel file.")
        return

    for sheet_name in sheet_names:
        movies = fetch_movie_list_local(GHIB_FILE, sheet_name)
        if not movies:
            log_debug(f"[INFO] Sheet '{sheet_name}' has no movies in col A.")
            continue

        # Make sure the JSON for this sheet exists
        json_file = ensure_url_json_exists(sheet_name)

        # Fill missing trailer URLs with the master cache approach
        fill_missing_urls_in_json_with_cache(json_file, movies, master_cache)

    print("\nAll sheets processed! Check 'trailer_debug.log' for logs.")

# -------------- RUN -------------- #
if __name__ == "__main__":
    fill_missing_urls_for_all_sheets_with_duplicates_tmdb()
    print("Done!")
