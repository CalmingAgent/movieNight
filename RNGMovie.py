from __future__ import annotations

import datetime
import importlib
import json
import os
import random
import re
import subprocess
import sys
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
import requests
from dotenv import load_dotenv
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from difflib import get_close_matches

import qtawesome as qta

try:
    from PySide6.QtCore import Qt, QUrl, QSize, Slot
    from PySide6.QtGui import QAction, QIcon, QColor, QPalette, QDesktopServices, QPixmap, QPainter, QFont, QColor   
    from PySide6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QListWidget, QListWidgetItem, QLabel, QStackedWidget, QPushButton,
        QLineEdit, QSplitter, QScrollArea, QTableWidget, QTableWidgetItem,
        QDialog, QCheckBox, QDialogButtonBox, QMessageBox, QGraphicsDropShadowEffect
    )
except ModuleNotFoundError as exc:
    sys.stderr.write(
        "PySide6 is not installed. GUI will not run.\n"
        "Install with:  pip install PySide6\n"
    )
    raise exc


# ────────────────────────── Configuration ──────────────────────────
BASE_DIR = Path(__file__).resolve().parent

# file / folder paths
ENV_PATH            = BASE_DIR / "secret.env"
TRAILER_FOLDER      = BASE_DIR / "Video_Trailers"
NUMBERS_FOLDER      = BASE_DIR / "Numbers"
GHIBLI_SHEET_PATH   = BASE_DIR / "ghib.xlsx"
UNDER_REVIEW_PATH   = BASE_DIR / "underReviewURLs.json"
SERVICE_ACCOUNT_KEY = BASE_DIR / "service_secret.json"         # not used directly here
CLIENT_SECRET_PATH  = BASE_DIR / "client_secret.json"
USER_TOKEN_PATH     = BASE_DIR / "youtube_token.pickle"
LOG_PATH            = BASE_DIR / "trailer_debug.log"
AUTO_UPDATE_SCRIPT  = BASE_DIR / "autoUpdate.py"

# constants
ICON                = lambda name: QIcon(str(BASE_DIR / "icons" / f"{name}.svg"))
ACCENT_COLOR        = "#3b82f6"
YOUTUBE_SEARCH_URL  = "https://www.googleapis.com/youtube/v3/search"
DRIVE_SCOPES        = ["https://www.googleapis.com/auth/drive.readonly"]
YOUTUBE_SCOPES      = ["https://www.googleapis.com/auth/youtube"]

# env vars
load_dotenv(ENV_PATH)
SPREADSHEET_ID  = os.getenv("SPREADSHEET_ID")
YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY")
if not SPREADSHEET_ID or not YOUTUBE_API_KEY:
    raise EnvironmentError("Missing SPREADSHEET_ID or YOUTUBE_API_KEY in .env")


# ────────────────────────── Logging helper ─────────────────────────
def log_debug(message: str) -> None:
    """
    Append a timestamped debug line to LOG_PATH.

    Uses Path.write_text(..., append=True) when available (3.12+),
    otherwise falls back to explicit 'a' mode open – so it never crashes.
    """
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.datetime.now().isoformat(timespec="seconds")
    log_line = f"[{timestamp}] {message}\n"

    try:
        # Python ≥ 3.12
        LOG_PATH.write_text(log_line, encoding="utf-8", append=True)  # type: ignore[arg-type]
    except TypeError:
        # Older Python – no 'append' kwarg
        with LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(log_line)

# ────────────────────────── Tiny utilities ─────────────────────────
def normalize(text: str) -> str:
    """Lower-case, strip, and drop non-alphanumerics (for fuzzy keys)."""
    return re.sub(r"[^a-z0-9]", "", text.lower().strip())


def fuzzy_match(target: str, candidates: List[str], cutoff: float = 0.8) -> Optional[str]:
    """Return best fuzzy match or None."""
    matches = get_close_matches(target, candidates, n=1, cutoff=cutoff)
    return matches[0] if matches else None


def movie_probability(title: str) -> float:
    """Pull probability from a separate `probability.py` (if present)."""
    try:
        return float(importlib.import_module("probability").get_prob(title))
    except Exception:
        return 0.0
    
def make_number_pixmap(number: int,
                       size: int = 96,
                       fg_color: str = "#ffffff",
                       bg_color: str = "#1f1f1f",
                       border_color: str = "#3b82f6") -> QPixmap:

    pixmap = QPixmap(size, size)
    pixmap.fill(QColor(bg_color))

    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.Antialiasing)

    # draw border square
    pen = painter.pen()
    pen.setWidth(4)
    pen.setColor(QColor(border_color))
    painter.setPen(pen)
    painter.drawRoundedRect(2, 2, size - 4, size - 4, 10, 10)

    # draw number
    font = QFont("Arial", int(size * 0.55), QFont.Bold)
    painter.setFont(font)
    painter.setPen(QColor(fg_color))
    painter.drawText(pixmap.rect(), Qt.AlignCenter, str(number))
    painter.end()

    return pixmap
# ───────────────────────── trailer-report helper ─────────────────────────
def report_trailer(movie_name: str, youtube_url: str | None) -> None:
    """
    Add/overwrite an entry in underReviewURLs.json and pop a confirmation.

    JSON is written pretty-printed (indented, one key per line) so it’s easy
    to read or edit by hand.
    """
    # 1) make sure the file exists
    if not UNDER_REVIEW_PATH.exists():
        UNDER_REVIEW_PATH.write_text("{}", encoding="utf-8")

    # 2) read existing contents (gracefully handle bad JSON)
    try:
        data: dict[str, str] = json.loads(UNDER_REVIEW_PATH.read_text("utf-8"))
    except json.JSONDecodeError:
        data = {}

    # 3) update / set the entry
    data[movie_name] = youtube_url or ""

    # 4) write it back – pretty JSON, Unix newlines, trailing newline
    UNDER_REVIEW_PATH.write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    # 5) log and notify
    log_debug(f"[REPORT] Marked “{movie_name}” → {youtube_url} as under review.")
    QMessageBox.information(
        None,
        "Trailer reported",
        f"Marked “{movie_name}” for manual review."
    )

# ────────────────────────── YouTube helpers ────────────────────────
def search_youtube_api(query: str) -> Optional[Tuple[str, str]]:
    """Return (url, video_title) for first short video result or None."""
    params = {
        "part": "snippet",
        "q": query,
        "key": YOUTUBE_API_KEY,
        "videoDuration": "short",
        "maxResults": 1,
        "type": "video",
    }
    try:
        response = requests.get(YOUTUBE_SEARCH_URL, params=params, timeout=10)
        data = response.json()
        if data.get("items"):
            first = data["items"][0]
            video_id = first["id"]["videoId"]
            title = first["snippet"]["title"]
            return f"https://www.youtube.com/watch?v={video_id}", title
    except Exception as exc:
        log_debug(f"YouTube search error: {exc}")
    return None


def locate_trailer(work_sheet: str, movie_title: str) -> Tuple[Optional[str], str, Optional[str]]:
    """
    Try local JSON first, then YouTube API.
    Returns (url, source, api_title) where source is 'json' or 'youtube'.
    """
    json_name = re.sub(r"\s+", "", work_sheet)
    urls_path = TRAILER_FOLDER / f"{json_name}Urls.json"

    if urls_path.exists():
        data = json.loads(urls_path.read_text(encoding="utf-8"))
        normalized = {normalize(k): v for k, v in data.items()}
        key = normalize(movie_title)
        url = normalized.get(key) or normalized.get(fuzzy_match(key, list(normalized)) or "")
        if url:
            return url, "json", None

    # fallback: YouTube
    api_result = search_youtube_api(f"{movie_title} official trailer")
    if api_result:
        url, api_title = api_result
        return url, "youtube", api_title

    return None, "", None


# ──────────────────────── OAuth & playlist helpers ─────────────────
def get_youtube_service():
    """Return cached or freshly-authenticated youtube client."""
    creds = None
    if USER_TOKEN_PATH.exists():
        creds = pickle.loads(USER_TOKEN_PATH.read_bytes())

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(CLIENT_SECRET_PATH), YOUTUBE_SCOPES)
            creds = flow.run_local_server(port=0)
        USER_TOKEN_PATH.write_bytes(pickle.dumps(creds))

    return build("youtube", "v3", credentials=creds)


def create_youtube_playlist(title: str, video_ids: List[str]) -> Optional[str]:
    """Create an unlisted playlist populated with `video_ids`; return playlist URL or None."""
    try:
        youtube = get_youtube_service()
        playlist = youtube.playlists().insert(
            part="snippet,status",
            body={
                "snippet": {
                    "title": title,
                    "description": "Auto-generated by Movie Night",
                },
                "status": {"privacyStatus": "unlisted"},
            },
        ).execute()

        playlist_id = playlist["id"]
        for vid in video_ids:
            youtube.playlistItems().insert(
                part="snippet",
                body={
                    "snippet": {
                        "playlistId": playlist_id,
                        "resourceId": {"kind": "youtube#video", "videoId": vid},
                    }
                },
            ).execute()

        return f"https://www.youtube.com/playlist?list={playlist_id}"
    except Exception as exc:
        log_debug(f"YouTube playlist creation failed: {exc}")
        return None


# ────────────────────────── GUI widgets ───────────────────────────
class ReportDialog(QDialog):
    """Checkbox list of movies to mark as ‘bad trailer’."""

    def __init__(self, parent: QWidget, movies: List[str]) -> None:
        super().__init__(parent)
        self.setWindowTitle("Report trailers")

        layout = QVBoxLayout(self)
        self._boxes: List[QCheckBox] = [QCheckBox(title) for title in movies]
        for box in self._boxes:
            layout.addWidget(box)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_movies(self) -> List[str]:
        return [box.text() for box in self._boxes if box.isChecked()]


class PickerPage(QWidget):
    """Main picker UI (left controls, center movie list, right stats)."""

    DIRECTIONS = ["Clockwise", "Counter-Clockwise"]

    def __init__(self, main_window: "MainWindow") -> None:
        super().__init__()

        self.main_window = main_window
        self._build_layout()
        self._connect_signals()

        # show direction/number only after first Generate Movies
        self.direction_label.hide()
        self.number_label.hide()

    # ───────────── layout helpers ─────────────
    def _build_layout(self) -> None:
        outer = QHBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(20)        

        # left: controls
        control_layout = QVBoxLayout()
        self.attendee_input = QLineEdit(placeholderText="# attendees")
        self.sheet_input = QLineEdit(placeholderText="Sheet name")

        self.generate_button = QPushButton("Generate Movies")
        self.update_urls_button = QPushButton("Update URLs")

        for widget in (self.attendee_input, self.sheet_input,
                       self.generate_button, self.update_urls_button):
            control_layout.addWidget(widget)
        control_layout.addStretch()
        outer.addLayout(control_layout)

        # centre: scrollable movie list
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        card_frame = QWidget()
        card_frame.setObjectName("MovieCard")
        card_layout = QVBoxLayout(card_frame)
        card_layout.setContentsMargins(12, 12, 12, 12)
        # movie list lives inside the card
        self.movie_list_container = QWidget()
        self.movie_list_layout = QVBoxLayout(self.movie_list_container)
        self.movie_list_layout.setAlignment(Qt.AlignTop)
        card_layout.addWidget(self.movie_list_container)
        card_frame.setStyleSheet(
            "background:#272727; border-radius:12px;"
        )
        self.scroll_area.setWidget(self.movie_list_container)
        outer.addWidget(self.scroll_area, 2)

        # right: stats & direction
        right_layout = QVBoxLayout()

        self.stats_label = QLabel(alignment=Qt.AlignCenter)
        right_layout.addWidget(self.stats_label)

        # direction + number tiles
        self.direction_label = QLabel("", alignment=Qt.AlignCenter, objectName="DirectionTile")
        self.direction_label.setFixedSize(112, 112)
        self.direction_label.setStyleSheet("border: 2px solid #555; border-radius: 8px;")
        right_layout.addWidget(self.direction_label, alignment=Qt.AlignHCenter)

        self.number_label = QLabel("", alignment=Qt.AlignCenter, objectName="NumberTile")   
        self.number_label.setFixedSize(112, 112)
        self.number_label.setStyleSheet("border: 2px solid #555; border-radius: 8px;")
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(12)
        shadow.setColor(QColor(0, 0, 0, 160))   # semi-transparent black
        shadow.setOffset(0, 2)
        self.number_label.setGraphicsEffect(shadow)

        shadow2 = QGraphicsDropShadowEffect(self)
        shadow2.setBlurRadius(12)
        shadow2.setColor(QColor(0, 0, 0, 160))
        shadow2.setOffset(0, 2)
        self.direction_label.setGraphicsEffect(shadow2)
        right_layout.addWidget(self.number_label, alignment=Qt.AlignHCenter)

        self.report_button = QPushButton("Report Trailers")
        self.report_button.setEnabled(False)
        right_layout.addWidget(self.report_button)
        right_layout.addStretch()

        outer.addLayout(right_layout)

    def _connect_signals(self) -> None:
        self.generate_button.clicked.connect(self.main_window.generate_movies)
        self.update_urls_button.clicked.connect(self.main_window.update_urls)
        self.report_button.clicked.connect(self._open_report_dialog)

        # pressing Enter in either LineEdit triggers generate
        self.attendee_input.returnPressed.connect(self.main_window.generate_movies)
        self.sheet_input.returnPressed.connect(self.main_window.generate_movies)

    # ───────────── public API ─────────────
    def display_movies(self, movies: List[str], trailer_lookup: dict[str, str]) -> None:
        # clear old widgets
        while (item := self.movie_list_layout.takeAt(0)):
            item.widget().deleteLater()

        self.current_movies = movies
        self.current_lookup = trailer_lookup

        for title in movies:
            url = trailer_lookup.get(title, "")
            prob_val = movie_probability(title)
            prob_str = f"{prob_val:.2f}"

            # probability colour tiers
            if   prob_val >= 0.7: pill_bg = "#2ecc71"      # green
            elif prob_val >= 0.4: pill_bg = "#f1c40f"      # yellow
            else               : pill_bg = ACCENT_COLOR    # blue

            # probability pill as its own QLabel
            pill = QLabel(prob_str)
            pill.setObjectName("")          # no id
            pill.setProperty("class", "prob-pill")
            pill.setStyleSheet(f"background:{pill_bg};")

            # title label (link if URL)
            if url:
                title_html = f'<a href="{url}" style="text-decoration:none">{title}</a>'
            else:
                title_html = f"<span style='color:#888'>{title}</span>"
            
            title_lbl = QLabel(title_html)
            title_lbl.setTextFormat(Qt.RichText)
            title_lbl.setTextInteractionFlags(Qt.TextBrowserInteraction)
            title_lbl.setOpenExternalLinks(False)
            title_lbl.linkActivated.connect(lambda _, link=url: QDesktopServices.openUrl(QUrl(link)))

            # row container
            row = QHBoxLayout()
            row.addWidget(title_lbl)
            row.addStretch()
            row.addWidget(pill)

            wrapper = QWidget()
            wrapper.setLayout(row)
            self.movie_list_layout.addWidget(wrapper)
            self.movie_list_layout.setSpacing(8)

        self.report_button.setEnabled(bool(movies))

        # random direction + number
        # inside PickerPage.display_movies or wherever you pick direction/number
        direction = random.choice(self.DIRECTIONS)                    # "Clockwise" / "Counter-Clockwise"
        icon_map = {
            "Clockwise": "arrow-clockwise",
            "Counter-Clockwise": "arrow-counterclockwise",
        }
        icon_name = icon_map[direction]                               # e.g. "arrow-right-circle"

        number = random.randint(1, (int(self.attendee_input.text())))

        # build pixmaps
        icon_pixmap   = ICON(icon_name).pixmap(96, 96)                # 96×96 px (tweak size if needed)
        number_pixmap = make_number_pixmap(
                        number,
                        bg_color="rgba(59,130,246,0.25)",   # 25 % opacity accent
                        fg_color="#ffffff"
                        ) 

        # push to labels
        self.direction_label.setPixmap(icon_pixmap)
        self.number_label.setPixmap(number_pixmap)

        # optional: let Qt scale when label is resized
        self.direction_label.setScaledContents(True)
        self.number_label.setScaledContents(True)

        self.direction_label.show()
        self.number_label.show()

    # ───────────── internal slots ─────────────
    @Slot()
    def _open_report_dialog(self) -> None:
        dialog = ReportDialog(self, self.current_movies)
        if dialog.exec() == QDialog.Accepted:
            for title in dialog.selected_movies():
                report_trailer(title, self.current_lookup.get(title, ""))
            QMessageBox.information(self, "Reported", "Thanks for the feedback!")


class StatsPage(QWidget):
    """Placeholder stats page."""
    def __init__(self) -> None:
        super().__init__()
        self.table = QTableWidget(0, 3)  # TODO: implement real stats


# ────────────────────────── Main Window ───────────────────────────
class MainWindow(QMainWindow):

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Movie Night")
        self.resize(960, 600)

        self.picker_page = PickerPage(self)
        self.stats_page = StatsPage()

        # navigation list
        self.nav_list = QListWidget()
        self.nav_list.setFixedWidth(170)
        self.nav_list.addItem(QListWidgetItem(ICON("grid"), "RNJesus"))
        row = self.nav_list.count()                # index of the item you just appended
        item = self.nav_list.item(row - 1)         # last added item
        item.setTextAlignment(Qt.AlignHCenter)
        self.nav_list.addItem(QListWidgetItem(ICON("bar-chart-line"), "  Movie Stats"))
        row = self.nav_list.count()                # index of the item you just appended
        item = self.nav_list.item(row - 1)         # last added item
        item.setTextAlignment(Qt.AlignHCenter)


        # stacked pages
        self.pages = QStackedWidget()
        self.pages.addWidget(self.picker_page)
        self.pages.addWidget(self.stats_page)

        self.nav_list.currentRowChanged.connect(self.pages.setCurrentIndex)

        splitter = QSplitter()
        splitter.addWidget(self.nav_list)
        splitter.addWidget(self.pages)
        splitter.setStretchFactor(1, 1)
        self.setCentralWidget(splitter)

        # toolbar
        toolbar = self.addToolBar("Main")
        reroll_action = QAction(ICON("refresh"), "Re-roll", self)
        reroll_action.setShortcut("Ctrl+R")
        reroll_action.triggered.connect(self.generate_movies)
        toolbar.addAction(reroll_action)

    # ─────────────────── public slots ───────────────────
    @Slot()
    def update_urls(self) -> None:
        subprocess.run(["python", AUTO_UPDATE_SCRIPT], check=False)

    @Slot()
    def generate_movies(self) -> None:
        """Validate input, pick movies, build playlist, update UI."""
        try:
            attendee_count = int(self.picker_page.attendee_input.text().strip())
            if attendee_count <= 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "Error", "Enter a positive attendee count.")
            return

        sheet_name_raw = self.picker_page.sheet_input.text().strip()
        if not sheet_name_raw:
            QMessageBox.warning(self, "Error", "Sheet name?")
            return
        if not GHIBLI_SHEET_PATH.exists():
            QMessageBox.warning(self, "Error", "Run Update URLs first.")
            return

        workbook = openpyxl.load_workbook(GHIBLI_SHEET_PATH, read_only=True)
        sheet_map = {normalize(name): name for name in workbook.sheetnames}
        chosen_sheet = (
            sheet_map.get(normalize(sheet_name_raw))
            or sheet_map.get(fuzzy_match(normalize(sheet_name_raw), list(sheet_map)))
        )
        if not chosen_sheet:
            QMessageBox.warning(self, "Error", "Sheet not found.")
            return

        movie_titles = [
            row[0] for row in workbook[chosen_sheet].iter_rows(min_row=1, max_col=1, values_only=True)
            if row[0]
        ]
        if attendee_count + 1 > len(movie_titles):
            QMessageBox.warning(self, "Error", "Not enough movies.")
            return

        chosen_movies = random.sample(movie_titles, attendee_count + 1)
        trailer_lookup = {
            title: locate_trailer(chosen_sheet, title)[0] for title in chosen_movies
        }

        # optional: build YouTube playlist from found trailers
        video_ids = [
            url.split("v=")[-1].split("&")[0]
            for url in trailer_lookup.values()
            if url and "watch?v=" in url
        ]
        if video_ids:
            playlist_url = create_youtube_playlist(
                f"Movie Night {datetime.date.today()}", video_ids
            )
            if playlist_url:
                QDesktopServices.openUrl(QUrl(playlist_url))

        # push to UI
        self.picker_page.display_movies(chosen_movies, trailer_lookup)


# ────────────────────────── Dark theme ────────────────────────────
def apply_dark_palette(app: QApplication) -> None:
    """Dark Fusion palette with brand accent."""
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor("#202124"))
    palette.setColor(QPalette.WindowText, Qt.white)
    palette.setColor(QPalette.Base, QColor("#2b2c2e"))
    palette.setColor(QPalette.AlternateBase, QColor("#323336"))
    palette.setColor(QPalette.Button, QColor("#2d2e30"))
    palette.setColor(QPalette.ButtonText, Qt.white)
    palette.setColor(QPalette.Text, Qt.white)
    palette.setColor(QPalette.Link, QColor(ACCENT_COLOR))
    palette.setColor(QPalette.Highlight, QColor(ACCENT_COLOR))
    palette.setColor(QPalette.HighlightedText, Qt.white)
    app.setStyle("Fusion")
    app.setPalette(palette)


# ────────────────────────── Entry point ───────────────────────────
def main() -> None:
    app = QApplication([])
    apply_dark_palette(app)
    app.setStyleSheet("""
    /* round the square tiles and give them a shadow */
    QLabel#DirectionTile, QLabel#NumberTile {
        border        : 2px solid #555;
        border-radius : 10px;
        background    : #1f1f1f;
        padding       : 4px;
        box-shadow    : 0 2px 6px rgba(0,0,0,.45);
    }

    /* nicer pill for probability */
    QLabel.prob-pill {
        color         : #111;
        border-radius : 4px;
        padding       : 1px 5px;
        font-size     : 11px;
    }

    /* movie link colour */
    a { color: #F8CC6A; }

    /* nav hover + selected tint */
    QListWidget::item:hover { background : #333; }
    QListWidget::item:selected { background : #444; }

    QPushButton:hover { background : #404040; }
    """)
    MainWindow().show()
    app.exec()


if __name__ == "__main__":
    main()
